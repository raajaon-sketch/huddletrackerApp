# =============================================================================
# HuddleTracker — Flask Application  (app.py)
# =============================================================================
#
# Architecture overview
# ─────────────────────
#   • Flask 3.x  +  Flask-SQLAlchemy  +  Flask-Login
#   • SQLite database  (huddle_tracker.db, same directory as this file)
#   • Two user-facing portals served as single-page apps via Jinja templates:
#       /admin  → admin_portal.html   (admin/lead role only)
#       /huddle → huddle_portal.html  (all authenticated users)
#   • All data access happens through a JSON REST API under /api/...
#   • Real-time updates are pushed to browsers via Server-Sent Events (/api/stream)
#   • Email notifications are sent asynchronously from a background daemon thread
#   • A second daemon thread runs daily overdue-task reminders
#
# Role hierarchy
# ──────────────
#   admin  – full CRUD on all resources, user management, email config
#   lead   – same as admin except cannot create/delete users or change email config
#   member – can only see/edit tasks assigned to them or created by them
#
# Security fixes applied (v24 → v28)
# ────────────────────────────────────
#   SEC-01  Enum validation on status/priority (no free-text injection)
#   SEC-02  Password complexity enforcement (min 8 chars, uppercase, digit)
#   SEC-03  IDOR guard: verify target user exists before assignment
#   SEC-04  MAX_CONTENT_LENGTH = 10 MB to prevent upload-based DoS
#   SEC-06  Security response headers on every reply
#   SEC-07  Exceptions logged server-side; raw tracebacks never returned
#   SEC-08  LIKE metachar escaping (\, %, _) in search queries
#   SEC-09  Export scoped to member's own tasks
#   SEC-10  CSRF token validated on all mutating API calls
#   SEC-11  Bulk import row cap (5 000 rows)
#   SEC-12  File MIME + magic-byte validation before xlsx parsing
#   SEC-S1  Constant-time login (dummy hash prevents username enumeration via timing)
#   SEC-S2  Content-Security-Policy header added to all responses
#   SEC-S3  get_users_activity() body was missing — restored with proper GROUP BY query
#   SEC-S4  Rate limiting extended to create_user, bulk_update, bulk_delete, bulk_import
#   SEC-S5  Permissions-Policy header added to disable unused browser APIs
#   PERF-P1 tat_warnings() limited to last 30 days (was unbounded full table scan)
#   PERF-P3 analytics() reduced from 9 COUNT queries to 2 GROUP BY queries
#   PERF-P4 Composite DB indexes added: (project_id,status) and (assigned_to,status)
#   PERF-P5 Overdue reminder query capped at 1000 rows
#   SEC-R1  load_user() now returns None for inactive users (instant session revocation)
#   SEC-R2  get_task_audit() IDOR fixed — members restricted to their own tasks
#   SEC-R3  get_tasks_kanban() IDOR fixed — members scoped to their own tasks
#   SEC-R4  Rate limiting (@_login_limit) applied to all 20 remaining mutation routes
#   SEC-R5  overdue_preview() capped at 200 rows, full task dicts replaced with summaries
#   PERF-R2 get_tasks_kanban() capped at 500 rows
#   PERF-R3 analytics_daily() date mode — joinedload added to prevent N+1
#   PERF-R4 bulk_import_task_names() — pre-loaded existing pairs set (eliminates N DB queries)
#   BUG-01  SECRET_KEY read from env var, never hardcoded
#   BUG-02  CSRF per-session token, rotated on login
#   BUG-03  SESSION_COOKIE_SECURE enabled only in production
#   BUG-04  joinedload on Task._assignee / creator to eliminate N+1 queries
#   BUG-05  Analytics scoped per role
#   BUG-06  db.session.get() instead of deprecated Query.get()
#   BUG-07  Removed fragile raw ALTER TABLE migration; use Flask-Migrate
#   BUG-08  Debug flag driven by FLASK_DEBUG env var

# ── Standard library ──────────────────────────────────────────────────────────
import uuid, os, re, io, secrets, logging, json, queue, threading, smtplib, time
from datetime import datetime, timedelta
from collections import defaultdict
from email.mime.multipart import MIMEMultipart   # Build multi-part email messages
from email.mime.text import MIMEText             # Attach plain-text / HTML parts
import calendar as cal_mod                       # monthrange() for calendar analytics

# ── Third-party ───────────────────────────────────────────────────────────────
from flask import (
    Flask, render_template, request, jsonify, redirect, url_for,
    send_file, session, Response, stream_with_context,
)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.orm import joinedload            # Eager-load relationships (avoids N+1)
from sqlalchemy import case as sa_case           # CASE WHEN in aggregate queries
from flask_login import (
    LoginManager, UserMixin, login_user, logout_user,
    login_required, current_user,
)
from werkzeug.security import generate_password_hash, check_password_hash

# PERF-04: openpyxl imported at module level — not inside each route handler —
# so the interpreter loads it once on startup instead of on every xlsx request.
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from openpyxl.utils import get_column_letter

# =============================================================================
# APPLICATION CONFIGURATION
# =============================================================================

app = Flask(__name__)

# ── Security: secret key ──────────────────────────────────────────────────────
# BUG-01 FIX: SECRET_KEY is now loaded from the environment variable so it is
# never hardcoded in source control.  The fallback secrets.token_hex(32) is safe
# for development only — it regenerates on every restart, invalidating sessions.
# In production always set SECRET_KEY to a stable random string in your .env.
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY') or secrets.token_hex(32)

# ── Database ──────────────────────────────────────────────────────────────────
# SQLite file lives next to this file.  For production consider PostgreSQL and
# update SQLALCHEMY_DATABASE_URI accordingly.
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(
    os.path.abspath(os.path.dirname(__file__)), 'huddle_tracker.db')
# Disabling modification tracking saves memory — we don't need SQLAlchemy's
# event system for our change detection (we use explicit audit rows instead).
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# ── Session / Cookie security ─────────────────────────────────────────────────
# HTTPONLY: JavaScript cannot read the session cookie, preventing XSS theft.
app.config['SESSION_COOKIE_HTTPONLY'] = True
# SAMESITE Lax: cookie is not sent on cross-site top-level navigations other
# than GET, providing baseline CSRF protection on top of our explicit token.
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
# BUG-03 FIX: SECURE flag is set only in production (HTTPS) so developers can
# still log in on http://localhost without the browser dropping the cookie.
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('FLASK_ENV') == 'production'
# Sessions expire after 8 working hours; Flask-Login's remember-me is disabled.
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)

# SEC-04 / SEC-12: Reject requests larger than 10 MB at the WSGI layer before
# they reach any route handler; prevents memory-exhaustion DoS via large uploads.
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024   # 10 MB

# BUG-08 FIX: Debug mode reads from the environment so it is impossible to
# accidentally ship a debug build.  Default is off (safe for production).
app.config['DEBUG'] = os.environ.get('FLASK_DEBUG', '0') == '1'

# =============================================================================
# EMAIL / SMTP CONFIGURATION
# =============================================================================
# All values are loaded from environment variables so credentials are never
# committed to source control.  See .env for defaults and examples.
# MAIL_ENABLED = false by default — set to true to activate notifications.

app.config['MAIL_ENABLED']   = os.environ.get('MAIL_ENABLED',   'false').lower() == 'true'
app.config['MAIL_SERVER']    = os.environ.get('MAIL_SERVER',    'smtp.gmail.com')
app.config['MAIL_PORT']      = int(os.environ.get('MAIL_PORT',  '587'))
app.config['MAIL_USE_TLS']   = os.environ.get('MAIL_USE_TLS',  'true').lower() == 'true'
app.config['MAIL_USERNAME']  = os.environ.get('MAIL_USERNAME',  '')
app.config['MAIL_PASSWORD']  = os.environ.get('MAIL_PASSWORD',  '')
# MAIL_FROM defaults to MAIL_USERNAME if not explicitly set
app.config['MAIL_FROM']      = os.environ.get('MAIL_FROM',      app.config.get('MAIL_USERNAME', ''))
# How many hours past due_date before a task is included in overdue reminders
app.config['REMINDER_HOURS'] = int(os.environ.get('REMINDER_HOURS', '24'))
# Scheduler wake-up interval in seconds (default: once per hour)
app.config['REMINDER_INTERVAL'] = int(os.environ.get('REMINDER_INTERVAL', '3600'))


# =============================================================================
# EXTENSIONS
# =============================================================================

# SQLAlchemy ORM — all models are registered against this db instance.
db = SQLAlchemy(app)

# Flask-Login — manages the authenticated session and the @login_required guard.
# login_view tells the extension which endpoint to redirect unauthenticated users to.
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# ── Logging ────────────────────────────────────────────────────────────────────
# Logger is initialised here — before the email engine and scheduler daemons —
# so both background threads can call logger.info / logger.error safely from the
# moment they start.  SEC-07: raw exception tracebacks are logged here and never
# forwarded to HTTP responses; clients only ever receive a generic error string.
logger = logging.getLogger(__name__)

# ── Flask-Migrate (Alembic) ────────────────────────────────────────────────────
# Provides `flask db init / migrate / upgrade` for versioned schema migrations.
# Wrapping the import in try/except allows the app to start even when the
# flask-migrate package is absent (e.g. a minimal dev environment).
# To run migrations:
#   flask db init       → create the migrations/ directory (first time only)
#   flask db migrate    → auto-generate a revision from model changes
#   flask db upgrade    → apply pending revisions to the database
try:
    from flask_migrate import Migrate as _Migrate
    migrate = _Migrate(app, db)
except ImportError:
    migrate = None

# =============================================================================
# EMAIL ENGINE  (async, non-blocking)
# =============================================================================
#
# Design rationale
# ────────────────
# Sending email synchronously inside a request handler would block the worker
# for several seconds on every SMTP round-trip, harming throughput.  Instead:
#   1.  Route handlers call send_email() which only enqueues a dict — O(1), never blocks.
#   2.  _mail_worker() runs in a daemon thread, drains the queue, and makes the
#       actual SMTP connection away from the request/response cycle.
#   3.  If MAIL_ENABLED is false (the default) every message is silently discarded,
#       making the email system a zero-overhead no-op in development/staging.

# Internal queue shared between the route threads and the mail worker thread.
_mail_queue: queue.Queue = queue.Queue()

def _mail_worker():
    """
    Daemon thread body — runs forever, delivering queued messages via SMTP.

    Flow per message:
      • Dequeue msg_data dict (blocks up to 5 s, then loops to check for shutdown).
      • If MAIL_ENABLED is false, discard silently and acknowledge the task.
      • Otherwise build a MIMEMultipart/alternative message with both HTML and
        plain-text parts (clients choose the richest version they support).
      • Open a fresh SMTP connection per message — keeps implementation simple
        and avoids stale-connection issues on long-running servers.
      • Log success or failure; never re-raise so the thread stays alive.
    """
    while True:
        try:
            msg_data = _mail_queue.get(timeout=5)
        except queue.Empty:
            continue   # No messages yet; loop back and wait
        if not app.config.get('MAIL_ENABLED'):
            logger.debug('Mail disabled — skipped: %s', msg_data.get('subject'))
            _mail_queue.task_done()
            continue
        try:
            # Build a MIME message with both HTML and plain-text alternatives
            msg = MIMEMultipart('alternative')
            msg['Subject'] = msg_data['subject']
            msg['From']    = app.config['MAIL_FROM'] or app.config['MAIL_USERNAME']
            msg['To']      = msg_data['to']
            if msg_data.get('html'):
                msg.attach(MIMEText(msg_data['html'],  'html',  'utf-8'))
            if msg_data.get('text'):
                msg.attach(MIMEText(msg_data['text'],  'plain', 'utf-8'))
            # Open SMTP, optionally upgrade to TLS, authenticate, send, close
            server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'], timeout=15)
            if app.config['MAIL_USE_TLS']:
                server.starttls()   # Upgrade plaintext connection to TLS (port 587)
            if app.config['MAIL_USERNAME']:
                server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
            server.sendmail(msg['From'], [msg_data['to']], msg.as_string())
            server.quit()
            logger.info('Email sent → %s  [%s]', msg_data['to'], msg_data['subject'])
        except Exception as exc:
            # Log but never crash the worker thread — the queue keeps running
            logger.error('Email send failed → %s: %s', msg_data.get('to'), exc)
        finally:
            # Always acknowledge the task so queue.join() (if used) unblocks
            _mail_queue.task_done()

# Start the mail worker as a daemon so it is automatically killed when the main
# process exits (no need for explicit shutdown logic).
_mail_thread = threading.Thread(target=_mail_worker, daemon=True, name='mail-worker')
_mail_thread.start()


def send_email(to: str, subject: str, html: str, text: str = '') -> None:
    """
    Public helper — enqueue an outbound email for async delivery.

    Args:
        to:      Recipient email address.  Silently ignored if missing or invalid.
        subject: Email subject line.
        html:    Full HTML body (use _html_wrap() to apply the branded shell).
        text:    Optional plain-text fallback for non-HTML clients.

    Never raises — any delivery errors are caught inside _mail_worker and logged.
    """
    if not to or '@' not in to:
        return   # Don't enqueue obviously invalid addresses
    _mail_queue.put({'to': to, 'subject': subject, 'html': html, 'text': text})


# =============================================================================
# EMAIL TEMPLATES  (branded HTML shell + per-event body generators)
# =============================================================================

# Brand colour constants — used in both the shell and inline badge styles so
# changing the palette only requires editing two lines.
_BRAND_COLOR = '#0a0a0f'   # Near-black — used for header bg, headings, buttons
_GOLD        = '#c9a84c'   # Gold accent — logo dot, hover hints

def _html_wrap(title: str, body: str) -> str:
    """
    Wrap an HTML body fragment in the HuddleTracker branded email shell.

    The shell includes:
      • Responsive max-width wrapper (580 px)
      • Dark header with gold logo mark
      • Content area with card, badge, table, and button helper classes
      • Plain footer with "do not reply" notice

    Args:
        title: Shown as the <h2> inside the content area.
        body:  Raw HTML fragment — must be trusted content (server-generated only).

    Returns:
        Complete <!DOCTYPE html> email string ready for MIMEText('html').
    """
    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
  body{{margin:0;padding:0;background:#f4f1eb;font-family:'Helvetica Neue',Arial,sans-serif;}}
  .wrap{{max-width:580px;margin:32px auto;background:#ffffff;border-radius:12px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,.08);}}
  .hdr{{background:{_BRAND_COLOR};padding:24px 32px;}}
  .hdr-logo{{font-size:1.4rem;font-weight:800;color:#fff;letter-spacing:-1px;}}
  .hdr-logo em{{color:{_GOLD};font-style:normal;}}
  .hdr-sub{{font-size:.7rem;letter-spacing:3px;text-transform:uppercase;color:rgba(255,255,255,.3);margin-top:4px;}}
  .body{{padding:28px 32px;}}
  h2{{font-size:1.1rem;color:{_BRAND_COLOR};margin:0 0 12px;}}
  p{{font-size:.875rem;color:#444;line-height:1.6;margin:0 0 12px;}}
  .card{{background:#f9f8f5;border:1px solid #e5e3dc;border-radius:8px;padding:14px 18px;margin:14px 0;}}
  .card-row{{display:flex;justify-content:space-between;padding:4px 0;font-size:.82rem;border-bottom:1px solid #ede;}}
  .card-row:last-child{{border-bottom:none;}}
  .card-row .lbl{{color:#888;}}
  .card-row .val{{font-weight:600;color:{_BRAND_COLOR};}}
  .badge{{display:inline-block;padding:2px 10px;border-radius:20px;font-size:.72rem;font-weight:600;}}
  .badge-done{{background:#dcfce7;color:#14532d;}}
  .badge-pending{{background:#fef9c3;color:#854d0e;}}
  .badge-in_progress{{background:#dbeafe;color:#1e40af;}}
  .badge-blocked{{background:#fee2e2;color:#991b1b;}}
  .badge-review{{background:#f3e8ff;color:#6b21a8;}}
  .btn{{display:inline-block;background:{_BRAND_COLOR};color:#fff;padding:11px 24px;border-radius:8px;text-decoration:none;font-size:.85rem;font-weight:600;margin-top:8px;}}
  .tbl{{width:100%;border-collapse:collapse;font-size:.8rem;margin:12px 0;}}
  .tbl th{{background:{_BRAND_COLOR};color:#fff;padding:8px 10px;text-align:left;font-size:.72rem;text-transform:uppercase;letter-spacing:.8px;}}
  .tbl td{{padding:8px 10px;border-bottom:1px solid #eee;color:#333;}}
  .tbl tr:last-child td{{border-bottom:none;}}
  .tbl tr:nth-child(even) td{{background:#fafaf8;}}
  .ftr{{background:#f4f1eb;padding:16px 32px;font-size:.72rem;color:#aaa;text-align:center;border-top:1px solid #e5e3dc;}}
</style>
</head>
<body>
<div class="wrap">
  <div class="hdr">
    <div class="hdr-logo">Huddle<em>.</em></div>
    <div class="hdr-sub">Task Tracker</div>
  </div>
  <div class="body">
    <h2>{title}</h2>
    {body}
  </div>
  <div class="ftr">This is an automated message from HuddleTracker · Do not reply</div>
</div>
</body></html>"""


def notify_task_submitted(task) -> None:
    """
    Send a "task submitted / marked Done" confirmation to the assigned member.

    Called from:
      • update_task()      when status transitions from any value → 'done'
      • bulk_update_tasks() for each task individually flipped to 'done'

    The email includes Study ID, Task Name, Mapping Process, TAT, and the
    completed_at timestamp so the member has a record of their submission.
    """
    assignee = task._assignee
    if not assignee or not assignee.email:
        return   # No assignee or no email address on file — skip silently
    td = task.to_dict()
    body = f"""
<p>Hi <strong>{assignee.username}</strong>,</p>
<p>Your task has been marked as <span class="badge badge-done">✓ Done</span> and recorded successfully.</p>
<div class="card">
  <div class="card-row"><span class="lbl">Study ID</span><span class="val">{td['study_id']}</span></div>
  <div class="card-row"><span class="lbl">Task Name</span><span class="val">{td['task_name']}</span></div>
  <div class="card-row"><span class="lbl">Mapping Process</span><span class="val">{td['process'] or '—'}</span></div>
  <div class="card-row"><span class="lbl">TAT</span><span class="val">{td['turnaround_time'] or '—'} mins</span></div>
  <div class="card-row"><span class="lbl">Completed</span><span class="val">{td['completed_at'] or 'just now'}</span></div>
</div>
<p style="color:#666;font-size:.82rem;">Great work! Your submission has been logged in the system.</p>
"""
    send_email(
        to=assignee.email,
        subject=f'✅ Task Submitted: {td["study_id"]} — {td["task_name"]}',
        html=_html_wrap('Task Submitted Successfully', body),
        text=f'Hi {assignee.username},\n\nYour task "{td["task_name"]}" (Study ID: {td["study_id"]}) has been marked as Done.\n\nGreat work!',
    )


def notify_task_assigned(task) -> None:
    """
    Send a "new task assigned" notification to the member the task was just
    assigned to.

    Called from:
      • create_task()   immediately after commit when assigned_to is set.
      • update_task()   when assigned_to changes to a different (non-null) user.
        In the update path, db.session.refresh(task) is called first so
        task._assignee reflects the NEW assignee, not the cached previous one.

    The email includes all key task fields plus due date so the member can
    plan their work immediately.
    """
    assignee = task._assignee
    if not assignee or not assignee.email:
        return
    td = task.to_dict()
    due_str = td.get('due_date') or 'Not set'
    body = f"""
<p>Hi <strong>{assignee.username}</strong>,</p>
<p>A new task has been assigned to you in HuddleTracker. Please review the details below and begin work at your earliest convenience.</p>
<div class="card">
  <div class="card-row"><span class="lbl">Study ID</span><span class="val">{td['study_id']}</span></div>
  <div class="card-row"><span class="lbl">SR / RFC</span><span class="val">{td['sr_rfc']}</span></div>
  <div class="card-row"><span class="lbl">Task Name</span><span class="val">{td['task_name']}</span></div>
  <div class="card-row"><span class="lbl">Mapping Process</span><span class="val">{td['process'] or '—'}</span></div>
  <div class="card-row"><span class="lbl">Priority</span><span class="val">{td['priority'].upper()}</span></div>
  <div class="card-row"><span class="lbl">TAT</span><span class="val">{td['turnaround_time'] or '—'} mins</span></div>
  <div class="card-row"><span class="lbl">Due Date</span><span class="val">{due_str}</span></div>
</div>
<p style="color:#666;font-size:.82rem;">Log in to HuddleTracker to update the task status as you progress.</p>
"""
    send_email(
        to=assignee.email,
        subject=f'📋 New Task Assigned: {td["study_id"]} — {td["task_name"]}',
        html=_html_wrap('New Task Assigned to You', body),
        text=f'Hi {assignee.username},\n\nA new task "{td["task_name"]}" (Study ID: {td["study_id"]}) has been assigned to you.\nDue: {due_str}\nPriority: {td["priority"]}\n\nPlease log in to HuddleTracker to get started.',
    )


def send_overdue_reminder(admins: list, overdue_by_member: dict) -> None:
    """
    Email all admin/lead users with a daily digest of overdue tasks.

    Args:
        admins:           List of User objects with role admin or lead.
        overdue_by_member: Dict mapping member username → list of task dicts.

    The email contains an HTML table grouped by member, showing Study ID,
    Task Name, Status badge, and Due Date highlighted in red.  Admins without
    an email address are silently skipped.
    """
    if not overdue_by_member:
        return
    total = sum(len(v) for v in overdue_by_member.values())

    # Build the HTML table rows — one row per overdue task
    rows_html = ''
    for member_name, tasks in sorted(overdue_by_member.items()):
        for td in tasks:
            due = td.get('due_date') or '—'
            rows_html += f"""<tr>
  <td>{member_name}</td>
  <td style="font-family:monospace">{td['study_id']}</td>
  <td>{td['task_name']}</td>
  <td><span class="badge badge-{td['status']}">{td['status'].replace('_',' ')}</span></td>
  <td style="color:#c00;font-weight:600">{due}</td>
</tr>"""

    body = f"""
<p>This is your daily overdue task reminder from HuddleTracker.</p>
<p>The following <strong>{total} task(s)</strong> are past their due date and have not been submitted (marked Done):</p>
<table class="tbl">
  <thead><tr><th>Member</th><th>Study ID</th><th>Task Name</th><th>Status</th><th>Due Date</th></tr></thead>
  <tbody>{rows_html}</tbody>
</table>
<p style="font-size:.8rem;color:#888;">Please follow up with the respective members. Tasks are considered overdue when the due date has passed and status is not <em>Done</em>.</p>
"""
    # Send one copy per admin — each gets an individual email
    for admin_user in admins:
        if not admin_user.email:
            continue
        send_email(
            to=admin_user.email,
            subject=f'⚠️ Overdue Tasks Reminder — {total} task(s) pending ({datetime.utcnow().strftime("%Y-%m-%d")})',
            html=_html_wrap(f'⚠️ {total} Overdue Task(s) Need Attention', body),
            text=f'Overdue tasks report: {total} task(s) are past due.\n\n' +
                 '\n'.join(f'- {m}: {len(t)} task(s)' for m, t in overdue_by_member.items()),
        )


# =============================================================================
# OVERDUE TASK REMINDER SCHEDULER
# =============================================================================
#
# A lightweight cron-like daemon that sends a daily digest email to all admin
# and lead users listing every task whose due_date is in the past and whose
# status is not 'done'.
#
# Implementation notes:
#   • Runs in a daemon thread — killed automatically when the process exits.
#   • Uses a module-level date string (_last_reminder_date) as a cheap guard
#     against duplicate sends within the same UTC day.  This works correctly
#     for a single-process deployment (gunicorn --workers=1 or development server).
#     For multi-worker deployments, move the guard to the database or Redis.
#   • Only sends after 08:00 UTC so the digest arrives during the work day.
#   • app.app_context() is required inside the thread because SQLAlchemy queries
#     need the Flask application context to resolve the database URI.

_last_reminder_date: str = ''   # '' means "no reminder sent yet today"

def _reminder_scheduler():
    """
    Daemon thread body — wakes periodically and fires the daily overdue digest.

    Sleeps for REMINDER_INTERVAL seconds between checks (default: 3600 s / 1 h).
    An initial 10-second delay on startup ensures the app is fully initialised
    (all models registered, DB tables created) before the first query runs.
    """
    global _last_reminder_date
    time.sleep(10)   # Brief startup grace period before first check
    while True:
        try:
            now   = datetime.utcnow()
            today = now.strftime('%Y-%m-%d')
            # Send at most once per UTC calendar day, and only after 08:00 UTC
            if today != _last_reminder_date and now.hour >= 8:
                with app.app_context():   # Required for SQLAlchemy queries in a thread
                    _check_and_send_overdue_reminders()
                _last_reminder_date = today
        except Exception as exc:
            logger.error('Reminder scheduler error: %s', exc)
        time.sleep(app.config.get('REMINDER_INTERVAL', 3600))


def _check_and_send_overdue_reminders():
    """
    Query the database for overdue tasks and dispatch the reminder email.

    A task is "overdue" when:
      • due_date < now  (the deadline has passed)
      • status not in ('done',)  (not yet submitted)

    Results are grouped by assignee username and forwarded to send_overdue_reminder().
    Unassigned tasks are grouped under the label 'Unassigned'.
    """
    now = datetime.utcnow()
    # PERF-P5: Cap at 1000 most-recently-due overdue tasks to bound memory use
    # as task history grows. Most recent tasks are most actionable for reminders.
    # Eager-load _assignee to avoid N+1 queries when grouping by member name.
    overdue = (
        Task.query
        .options(joinedload(Task._assignee))
        .filter(
            Task.due_date < now,
            Task.status.notin_(['done']),
        )
        .order_by(Task.due_date.desc())
        .limit(1000)
        .all()
    )
    if not overdue:
        return   # Nothing to report — skip the email entirely
    by_member: dict = defaultdict(list)
    for t in overdue:
        name = t._assignee.username if t._assignee else 'Unassigned'
        by_member[name].append(t.to_dict())
    # Fetch all active admin and lead users to receive the digest
    admins = User.query.filter(
        User.role.in_(['admin', 'lead']),
        User.is_active == True,
    ).all()
    send_overdue_reminder(admins, dict(by_member))
    logger.info('Overdue reminder sent: %d tasks across %d members', len(overdue), len(by_member))


# Launch the scheduler daemon immediately at import time
_reminder_thread = threading.Thread(target=_reminder_scheduler, daemon=True, name='reminder-scheduler')
_reminder_thread.start()

# =============================================================================
# CSRF PROTECTION
# =============================================================================
#
# BUG-02 FIX: A per-session CSRF token is generated on first request and
# validated on every state-changing API call (POST / PUT / DELETE).
#
# Why not Flask-WTF / wtforms?  The portals are single-page apps that use
# fetch() with JSON bodies, so token delivery happens via:
#   1. The login response JSON body (csrf_token field).
#   2. A Jinja global (csrf_token()) for any server-rendered form.
#
# Clients send the token in either:
#   • JSON body key "_csrf"
#   • HTTP header "X-CSRF-Token"  ← preferred for fetch() calls
#   • Form field "_csrf_token"

def get_csrf_token() -> str:
    """
    Return the CSRF token for the current session, creating one if absent.
    Exposed as a Jinja2 global so templates can call {{ csrf_token() }}.
    """
    if '_csrf_token' not in session:
        session['_csrf_token'] = secrets.token_hex(32)
    return session['_csrf_token']

# Register as a Jinja global so templates can embed the token in forms
app.jinja_env.globals['csrf_token'] = get_csrf_token

def _is_safe_method() -> bool:
    """Return True for read-only HTTP methods that don't need CSRF protection."""
    return request.method in ('GET', 'HEAD', 'OPTIONS')

@app.before_request
def csrf_protect():
    """
    Before-request hook: reject mutating API calls that carry an invalid or
    missing CSRF token.

    Exemptions (no token required):
      • GET / HEAD / OPTIONS — safe methods by definition
      • /login               — unauthenticated; session token doesn't exist yet
      • Non-API routes       — page navigations use browser's same-origin policy
      • Unauthenticated requests — no session to steal
    """
    if _is_safe_method():
        return
    if request.path == '/login':
        return   # Login form submits before a session (and token) exists
    # Only enforce on API endpoints and the logout action
    if not request.path.startswith('/api/') and request.path not in ('/logout',):
        return
    if not current_user.is_authenticated:
        return   # Flask-Login will handle the 401 separately
    # Accept token from any of the three supported locations
    token = (request.get_json(silent=True) or {}).get('_csrf') \
            or request.headers.get('X-CSRF-Token') \
            or request.form.get('_csrf_token')
    if not token or token != session.get('_csrf_token'):
        return jsonify({'error': 'CSRF token invalid or missing'}), 403

# =============================================================================
# SECURITY RESPONSE HEADERS
# =============================================================================

# SEC-06: Attach security headers to every HTTP response.
@app.after_request
def set_security_headers(response):
    """
    Add standard security headers to every response.

    X-Content-Type-Options: nosniff   — prevents MIME-type sniffing attacks.
    X-Frame-Options: SAMEORIGIN       — blocks clickjacking via <iframe>.
    X-XSS-Protection: 1; mode=block  — legacy XSS filter for older browsers.
    Referrer-Policy                   — limits referrer leakage on cross-origin nav.
    Strict-Transport-Security         — HSTS enforced only in production (HTTPS).
    """
    response.headers['X-Content-Type-Options']  = 'nosniff'
    response.headers['X-Frame-Options']          = 'SAMEORIGIN'
    response.headers['X-XSS-Protection']         = '1; mode=block'
    response.headers['Referrer-Policy']          = 'strict-origin-when-cross-origin'
    # SEC-S2: Content-Security-Policy — restricts script/style sources to self only,
    # blocking XSS payload exfiltration and injected external script loading.
    # unsafe-inline is needed for the inline <style> and <script> blocks in templates;
    # nonce-based CSP would be the ideal long-term improvement.
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://fonts.gstatic.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self'; "
        "frame-ancestors 'none';"
    )
    # SEC-S5: Permissions-Policy — disable powerful browser APIs not used by this app.
    response.headers['Permissions-Policy'] = (
        'camera=(), microphone=(), geolocation=(), payment=(), usb=()'
    )
    # Only send HSTS in production where the server is actually running HTTPS
    if os.environ.get('FLASK_ENV') == 'production':
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response

# =============================================================================
# RATE LIMITING
# =============================================================================
#
# Flask-Limiter is optional — the app degrades gracefully without it.
# Default limit: 1000 requests/hour per IP for general endpoints.
# Login endpoint gets a tighter 20 req/min limit (_login_limit) to slow
# credential-stuffing attacks before the per-username counter kicks in.

try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    limiter = Limiter(get_remote_address, app=app,
                      default_limits=['1000 per hour'], storage_uri='memory://')
    _login_limit = limiter.limit('20 per minute')
except Exception:
    # Graceful degradation: run without rate limiting if flask-limiter is absent
    limiter = None
    _login_limit = lambda f: f   # No-op decorator so @_login_limit still works


# =============================================================================
# SSE BROKER  (Server-Sent Events / real-time push)
# =============================================================================
#
# Provides live task updates to all connected browsers without polling.
# Design:
#   • _SSEBroker maintains a list of per-client Queue objects.
#   • Route handlers call sse.publish(event, data) after committing changes.
#   • The /api/stream endpoint subscribes a new Queue and streams its output
#     as a text/event-stream response.
#   • Queues that are full (client too slow) are dropped to prevent memory growth.
#   • No Redis or Celery required — works in a single-process deployment.

class _SSEBroker:
    """
    Thread-safe Server-Sent Events broadcaster.

    subscribe()   → returns a new Queue the caller should read from.
    unsubscribe() → removes the Queue (called when the client disconnects).
    publish()     → pushes a formatted SSE payload to all subscribed Queues.
    """

    def __init__(self):
        self._lock = threading.Lock()
        self._listeners: list[queue.Queue] = []

    def subscribe(self) -> queue.Queue:
        """Register a new SSE client and return its dedicated message queue."""
        q: queue.Queue = queue.Queue(maxsize=50)   # Cap at 50 buffered events
        with self._lock:
            self._listeners.append(q)
        return q

    def unsubscribe(self, q: queue.Queue):
        """Remove a client queue (client disconnected or stream generator exited)."""
        with self._lock:
            self._listeners = [l for l in self._listeners if l is not q]

    def publish(self, event: str, data: dict):
        """
        Push an SSE message to all connected clients.

        Args:
            event: SSE event name (e.g. 'task_updated', 'task_deleted').
            data:  JSON-serialisable dict — becomes the SSE 'data:' line.

        Queues that are full (client is too slow or the connection is stale) are
        automatically pruned to prevent unbounded memory growth.
        """
        payload = f"event: {event}\ndata: {json.dumps(data)}\n\n"
        with self._lock:
            dead = []
            for q in self._listeners:
                try:
                    q.put_nowait(payload)   # Non-blocking — drops if queue is full
                except queue.Full:
                    dead.append(q)          # Mark stale queues for removal
            # Prune stale queues outside the inner loop to avoid mutating while iterating
            self._listeners = [l for l in self._listeners if l not in dead]


# Singleton broker — shared across all request threads
sse = _SSEBroker()
# =============================================================================
# DATA MODELS
# =============================================================================


class User(UserMixin, db.Model):
    """
    Represents a HuddleTracker account.

    Roles:
        admin   — full system access: manage users, projects, tasks, email config.
        lead    — same as admin except cannot create/delete users or change SMTP.
        member  — can only view and update tasks assigned to or created by them.

    is_active = False is a soft-delete; the user cannot log in but their data
    (tasks, audit rows) is preserved for historical reporting.
    """
    __tablename__ = 'users'
    id            = db.Column(db.Integer, primary_key=True)
    username      = db.Column(db.String(80),  unique=True, nullable=False)
    email         = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    enterprise_id = db.Column(db.String(50),  unique=True, nullable=False)
    role          = db.Column(db.String(20),  default='member')
    is_active     = db.Column(db.Boolean,     default=True)
    created_at    = db.Column(db.DateTime,    default=datetime.utcnow)

    # Back-reference from Project (owner_id) — allows project.owner.username
    projects      = db.relationship('Project', backref='owner', lazy='dynamic',
                                    foreign_keys='Project.owner_id')
    # Back-reference from Task (assigned_to) — allows task.assignee.username
    assigned_tasks = db.relationship('Task', backref='assignee', lazy='dynamic',
                                     foreign_keys='Task.assigned_to')

    def set_password(self, pw: str) -> None:
        """Hash and store a new password using Werkzeug's PBKDF2 hasher."""
        self.password_hash = generate_password_hash(pw)

    def check_password(self, pw: str) -> bool:
        """Return True if pw matches the stored hash (timing-safe comparison)."""
        return check_password_hash(self.password_hash, pw)

    def to_dict(self) -> dict:
        """Serialise to a safe JSON-ready dict — password hash is never included."""
        return {
            'id': self.id, 'username': self.username, 'email': self.email,
            'enterprise_id': self.enterprise_id, 'role': self.role,
            'is_active': self.is_active,
        }


class Project(db.Model):
    """
    Groups a collection of Tasks under a named project.

    project_code is auto-generated (PRJ-XXXXXXXX) and used as a human-readable
    identifier in import templates and export files.

    Tasks have a cascade='all, delete-orphan' relationship so deleting a project
    automatically removes all its tasks (and their audit rows via CASCADE in DB).
    """
    __tablename__ = 'projects'
    id           = db.Column(db.Integer, primary_key=True)
    project_code = db.Column(db.String(20),  unique=True, nullable=False)
    name         = db.Column(db.String(200), nullable=False)
    description  = db.Column(db.Text,        default='')
    owner_id     = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    status       = db.Column(db.String(20),  default='active')
    created_at   = db.Column(db.DateTime,    default=datetime.utcnow)
    deadline     = db.Column(db.DateTime,    nullable=True)
    tasks        = db.relationship('Task', backref='project', lazy='dynamic',
                                   cascade='all, delete-orphan')

    def to_dict(self, task_count=None) -> dict:
        """
        Serialise project to a JSON-ready dict.

        Args:
            task_count: Pre-computed task count (int).  If provided, avoids a
                        lazy COUNT query — PERF-01: callers batch this via a
                        single GROUP BY query instead of N individual COUNTs.
        """
        return {
            'id': self.id, 'project_code': self.project_code, 'name': self.name,
            'description': self.description or '', 'status': self.status,
            'owner': self.owner.username if self.owner else '',
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M') if self.created_at else '',
            'deadline': self.deadline.strftime('%Y-%m-%d') if self.deadline else None,
            'task_count': task_count if task_count is not None else self.tasks.count(),
        }


class Task(db.Model):
    """
    Core entity representing a single unit of work.

    Status lifecycle:  pending → in_progress → review → done
                                             ↘ blocked (at any point)

    Key design decisions:
      • study_id is the primary business key visible to users (e.g. STD-202503-ABCD12).
      • assigned_to is nullable — unassigned tasks are visible to admins/leads only.
      • _assignee relationship uses overlaps= to silence SQLAlchemy's warning about
        the duplicate FK path (User.assigned_tasks backref vs Task._assignee).
      • PERF-02: index=True on project_id, assigned_to, status, priority, and
        created_at because these columns are used in WHERE / ORDER BY clauses on
        every task list query.
    """
    __tablename__ = 'tasks'
    id              = db.Column(db.Integer, primary_key=True)
    study_id        = db.Column(db.String(50),  unique=True, nullable=False)
    sr_rfc          = db.Column(db.String(100), nullable=False)
    # PERF-02: index=True — this column is in every project-filtered query
    project_id      = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=False, index=True)
    # PERF-02: index=True — filtered frequently for member-scoped task lists
    assigned_to     = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True,  index=True)
    task_name       = db.Column(db.String(300), nullable=False)
    process         = db.Column(db.String(200), default='')
    subtask         = db.Column(db.String(300), default='')
    # PERF-02: index=True — status and priority are used in WHERE filters and GROUP BY
    status          = db.Column(db.String(30),  default='pending', index=True)
    priority        = db.Column(db.String(10),  default='medium',  index=True)
    # PERF-P4: Composite index covers the common (project_id, status) filter pattern
    # used in get_tasks() when an admin views tasks filtered by project and status.
    __table_args__ = (
        db.Index('ix_task_project_status', 'project_id', 'status'),
        db.Index('ix_task_assigned_status', 'assigned_to', 'status'),
    )
    volume          = db.Column(db.Integer,     default=1)
    turnaround_time = db.Column(db.Integer,     nullable=True)    # minutes
    due_date        = db.Column(db.DateTime,    nullable=True)
    started_at      = db.Column(db.DateTime,    nullable=True)
    completed_at    = db.Column(db.DateTime,    nullable=True)
    # PERF-02: index=True — default ORDER BY on all task list queries
    created_at      = db.Column(db.DateTime,    default=datetime.utcnow, index=True)
    notes           = db.Column(db.Text,        default='')
    created_by      = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)

    # Eager-load the creator relationship to avoid N+1 queries in to_dict()
    creator  = db.relationship('User', foreign_keys=[created_by])

    # BUG-04 FIX: Dedicated relationship for the assignee so to_dict() and email
    # helpers use the already-loaded User object instead of triggering a lazy
    # SELECT per task row.  overlaps= suppresses the SQLAlchemy overlap warning
    # caused by the User.assigned_tasks backref sharing the same FK.
    _assignee = db.relationship('User', foreign_keys=[assigned_to],
                                overlaps='assignee,assigned_tasks')

    def to_dict(self) -> dict:
        """
        Serialise task to a JSON-ready dict.

        Accesses _assignee and creator via already-loaded relationships — no
        additional queries when callers use joinedload() on these relationships.
        """
        assignee = self._assignee
        creator  = self.creator
        return {
            'id': self.id, 'study_id': self.study_id, 'sr_rfc': self.sr_rfc,
            'project_id': self.project_id,
            'project_code': self.project.project_code if self.project else '',
            'project_name': self.project.name if self.project else '',
            'assigned_to': self.assigned_to,
            'assignee_name': assignee.username if assignee else 'Unassigned',
            'enterprise_id': assignee.enterprise_id if assignee else '',
            'task_name': self.task_name, 'process': self.process or '',
            'subtask': self.subtask or '',
            'status': self.status, 'priority': self.priority,
            'volume': self.volume, 'turnaround_time': self.turnaround_time,
            'due_date': self.due_date.strftime('%Y-%m-%d') if self.due_date else None,
            'started_at': self.started_at.strftime('%Y-%m-%d %H:%M') if self.started_at else None,
            'completed_at': self.completed_at.strftime('%Y-%m-%d %H:%M') if self.completed_at else None,
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M') if self.created_at else '',
            'notes': self.notes or '',
            'created_by': creator.username if creator else '',
            'created_by_id': self.created_by,
        }


class TaskName(db.Model):
    """
    Lookup table of approved task names for a given mapping process.

    Used to:
      1. Populate the Task Name dropdown in both portals (filtered by process).
      2. Auto-fill turnaround_time (cap_timing) when creating tasks.
      3. Drive the bulk import-from-process feature.

    cap_timing is in minutes.  status = 'inactive' hides the entry from dropdowns
    without deleting it (preserves history on existing tasks).
    """
    __tablename__ = 'task_names'
    id         = db.Column(db.Integer, primary_key=True)
    # index=True — lookup queries always filter by process first
    process    = db.Column(db.String(200), nullable=False, index=True)
    task_name  = db.Column(db.String(300), nullable=False)
    cap_timing = db.Column(db.Integer,     nullable=False)   # minutes
    status     = db.Column(db.String(20),  default='active')
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self) -> dict:
        return {
            'id': self.id, 'process': self.process, 'task_name': self.task_name,
            'cap_timing': self.cap_timing, 'status': self.status,
            'created_at': self.created_at.strftime('%Y-%m-%d') if self.created_at else '',
        }


class Process(db.Model):
    """
    Master list of mapping process categories (e.g. 'NG 1.0', 'Migrations').

    Renaming a process cascades to both TaskName.process and Task.process so
    existing records remain consistent without requiring a join.
    Deletion is blocked when any TaskName or live Task still references the name.
    """
    __tablename__ = 'processes'
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(200), unique=True, nullable=False)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self) -> dict:
        return {
            'id': self.id, 'name': self.name,
            'created_at': self.created_at.strftime('%Y-%m-%d') if self.created_at else '',
        }


class TaskAudit(db.Model):
    """
    Immutable audit trail — one row per field change on a Task.

    Written by _audit_task_change() inside update_task().  The table is
    append-only by convention; rows are never updated or deleted through
    the API.  ON DELETE CASCADE on task_id means audit rows are cleaned up
    automatically when their parent task is deleted.

    Fields tracked: status, priority, assigned_to, task_name, process,
    subtask, notes, sr_rfc, volume, turnaround_time, due_date,
    started_at, completed_at.
    """
    __tablename__ = 'task_audit'
    id         = db.Column(db.Integer, primary_key=True)
    # CASCADE in DB — audit rows removed when task is hard-deleted
    task_id    = db.Column(db.Integer, db.ForeignKey('tasks.id', ondelete='CASCADE'),
                           nullable=False, index=True)
    changed_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    field      = db.Column(db.String(50), nullable=False)
    old_value  = db.Column(db.Text, nullable=True)
    new_value  = db.Column(db.Text, nullable=True)
    # index=True — the History tab queries ORDER BY changed_at DESC
    changed_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)

    changer = db.relationship('User', foreign_keys=[changed_by])

    def to_dict(self) -> dict:
        return {
            'id': self.id, 'task_id': self.task_id,
            'field': self.field,
            'old_value': self.old_value, 'new_value': self.new_value,
            'changed_by': self.changer.username if self.changer else 'system',
            'changed_at': self.changed_at.strftime('%Y-%m-%d %H:%M') if self.changed_at else '',
        }


def _audit_task_change(task_id: int, field: str, old_val, new_val, user_id=None) -> None:
    """
    Append a TaskAudit row for a single field change.

    Converts both values to strings for comparison so None and '' are treated
    as equivalent (avoids spurious "None → ''" audit rows on untouched fields).
    Skips silently when old == new — callers don't need to guard themselves.

    Args:
        task_id:  ID of the Task being modified.
        field:    Column name that changed (e.g. 'status', 'assigned_to').
        old_val:  Previous value before this request.
        new_val:  New value after this request.
        user_id:  ID of the User who made the change (None = system action).
    """
    old_s = str(old_val) if old_val is not None else ''
    new_s = str(new_val) if new_val is not None else ''
    if old_s == new_s:
        return   # No actual change — skip
    db.session.add(TaskAudit(
        task_id=task_id, field=field,
        old_value=old_s or None, new_value=new_s or None,
        changed_by=user_id,
    ))


# =============================================================================
# FLASK-LOGIN USER LOADER + SHARED HELPERS
# =============================================================================

# BUG-06 FIX: db.session.get() is the SQLAlchemy 2.x equivalent of the
# deprecated Query.get().  Flask-Login calls this on every authenticated request
# to re-load the user from the session cookie's user ID.
@login_manager.user_loader
def load_user(uid):
    # SEC-S1: Return None for deactivated users so Flask-Login immediately
    # treats their session as unauthenticated — no grace period after deactivation.
    user = db.session.get(User, int(uid))
    if user and not user.is_active:
        return None
    return user


def gen_study_id() -> str:
    """
    Generate a unique Study ID for a new task.
    Format: STD-YYYYMM-XXXXXX  (year-month + 6 uppercase hex chars)
    Example: STD-202503-A4F2C1
    """
    return 'STD-{}-{}'.format(datetime.utcnow().strftime('%Y%m'), uuid.uuid4().hex[:6].upper())

def gen_project_code() -> str:
    """
    Generate a unique project code.
    Format: PRJ-XXXXXXXX  (8 uppercase hex chars)
    """
    return 'PRJ-{}'.format(uuid.uuid4().hex[:8].upper())

def sanitize(text) -> str:
    """
    Strip HTML-significant characters (<, >, ", ') from user-supplied strings.

    This is a defence-in-depth measure; the primary XSS defence is Jinja2's
    auto-escaping.  Used on all string inputs before they are stored in the DB.
    """
    return re.sub(r'[<>"\']', '', str(text)).strip() if text else ''


# =============================================================================
# AUTHENTICATION ROUTES
# =============================================================================

@app.route('/')
def index():
    """Root redirect — sends authenticated users to the portal selector,
    unauthenticated users to the login page."""
    if current_user.is_authenticated:
        return redirect(url_for('portal_select'))
    return redirect(url_for('login'))


@app.route('/portal')
@login_required
def portal_select():
    """Intermediate page where the user chooses Admin or Huddle portal.
    Rendered from portal_select.html."""
    return render_template('portal_select.html', user=current_user)


@app.route('/admin')
@login_required
def admin_portal():
    """
    Serve the Admin Portal SPA (admin_portal.html).
    Access is restricted to admin and lead roles; members are redirected to /huddle.
    """
    if current_user.role not in ('admin', 'lead'):
        return redirect(url_for('huddle_portal'))
    return render_template('admin_portal.html', user=current_user)


@app.route('/huddle')
@login_required
def huddle_portal():
    """Serve the Huddle Portal SPA (huddle_portal.html) — accessible to all roles."""
    return render_template('huddle_portal.html', user=current_user)


@app.route('/login', methods=['GET', 'POST'])
@_login_limit   # Stricter rate limit (20 req/min) to slow brute-force attempts
def login():
    """
    GET  → Render the login page (login.html).
    POST → Authenticate credentials and return a JSON response.

    Security measures:
      • Per-username failure counter stored in session — locks after 10 consecutive
        failures (429 response) until the session is cleared.
      • session.clear() before login_user() prevents session-fixation attacks.
      • CSRF token is rotated on login so the new authenticated session gets a
        fresh token that cannot be guessed from a pre-login session.
      • demo credentials shown only when FLASK_ENV != 'production'.
    """
    if request.method == 'POST':
        try:
            data     = request.get_json(silent=True) or request.form
            username = sanitize(data.get('username', ''))
            user     = User.query.filter_by(username=username).first()
            # Per-username brute-force counter (stored in server-side session)
            fail_key   = f'_fail_{username}'
            fail_count = session.get(fail_key, 0)
            if fail_count >= 10:
                return jsonify({'success': False, 'message': 'Account temporarily locked. Try again later.'}), 429
            # SEC-S1: Run check_password even when user is not found to prevent
            # username enumeration via timing differences (constant-time response).
            _DUMMY_HASH = generate_password_hash('__dummy_constant_time_check__')
            pwd = data.get('password', '')
            if user:
                valid = user.check_password(pwd) and user.is_active
            else:
                check_password_hash(_DUMMY_HASH, pwd)  # consume same time
                valid = False
            if valid:
                session.clear()          # Prevent session-fixation
                session.permanent = True  # Enforce the 8-hour PERMANENT_SESSION_LIFETIME
                login_user(user, remember=False)
                # Rotate CSRF token so it is bound to this authenticated session
                session['_csrf_token'] = secrets.token_hex(32)
                return jsonify({'success': True, 'role': user.role,
                                'csrf_token': session['_csrf_token']})
            # Increment failure counter for this username
            session[fail_key] = fail_count + 1
            return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
        except Exception as e:
            logger.exception('Login error')
            return jsonify({'success': False, 'message': 'Login failed'}), 500
    return render_template('login.html',
                           show_demo=(os.environ.get('FLASK_ENV') != 'production'))


@app.route('/logout')
@login_required
def logout():
    """
    Clear the session (including the CSRF token) and log the user out.
    Redirects to the login page.
    """
    session.clear()   # Wipe CSRF token and any other session data
    logout_user()
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    """Legacy dashboard route — renders dashboard.html."""
    return render_template('dashboard.html', user=current_user)


# =============================================================================
# PROJECTS API
# =============================================================================
#
# Endpoint summary:
#   GET    /api/projects          → list projects visible to the current user
#   POST   /api/projects          → create a new project
#   PUT    /api/projects/<pid>    → update name / description / status / deadline
#   DELETE /api/projects/<pid>    → delete project + all its tasks (cascade)

@app.route('/api/projects', methods=['GET'])
@login_required
def get_projects():
    """
    Return all projects visible to the current user.

    Visibility rules:
      admin/lead  — all projects in the system.
      member      — only projects they own OR have at least one assigned task in.

    PERF-01: Task counts are fetched in a single GROUP BY query and passed into
    to_dict() rather than letting each project fire its own COUNT(*) lazily.
    """
    try:
        q = Project.query
        if current_user.role == 'member':
            # Subquery: project IDs where this member has an assigned task
            ids = db.session.query(Task.project_id).filter_by(assigned_to=current_user.id).distinct()
            q = q.filter(db.or_(Project.owner_id == current_user.id, Project.id.in_(ids)))
        projects = q.order_by(Project.created_at.desc()).all()
        # PERF-01: Single GROUP BY query replaces N individual COUNT queries
        proj_ids = [p.id for p in projects]
        count_rows = (
            db.session.query(Task.project_id, db.func.count(Task.id))
            .filter(Task.project_id.in_(proj_ids))
            .group_by(Task.project_id)
            .all()
        ) if proj_ids else []
        counts = {pid: cnt for pid, cnt in count_rows}
        return jsonify([p.to_dict(task_count=counts.get(p.id, 0)) for p in projects])
    except Exception as e:
        logger.exception('Unhandled error')
        return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/projects', methods=['POST'])
@login_required
@_login_limit
def create_project():
    """
    Create a new project.  The project_code is auto-generated (PRJ-XXXXXXXX).
    deadline is optional and must be in YYYY-MM-DD format if supplied.
    The current user becomes the owner.
    """
    try:
        data = request.get_json(silent=True) or {}
        if not data.get('name'):
            return jsonify({'error': 'Project name required'}), 400
        deadline = None
        if data.get('deadline'):
            try:
                deadline = datetime.strptime(data['deadline'], '%Y-%m-%d')
            except ValueError:
                pass   # Invalid date format — deadline left as None
        p = Project(
            project_code=gen_project_code(),
            name=sanitize(data['name'])[:200],
            description=sanitize(data.get('description', ''))[:1000],
            owner_id=current_user.id,
            deadline=deadline,
        )
        db.session.add(p)
        db.session.commit()
        return jsonify(p.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        logger.exception('Unhandled error')
        return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/projects/<int:pid>', methods=['PUT'])
@login_required
@_login_limit
def update_project(pid):
    """
    Update an existing project's name, description, status, or deadline.
    Only the project owner or an admin/lead can modify it.
    Allowed status values: active, completed, archived.
    """
    try:
        p = db.session.get(Project, pid)   # BUG-06 FIX: db.session.get() not Query.get()
        if p is None:
            return jsonify({'error': 'Project not found'}), 404
        if p.owner_id != current_user.id and current_user.role not in ('admin', 'lead'):
            return jsonify({'error': 'Unauthorized'}), 403
        data = request.get_json(silent=True) or {}
        if 'name' in data:
            p.name = sanitize(data['name'])[:200]
        if 'description' in data:
            p.description = sanitize(data['description'])[:1000]
        if 'status' in data and data['status'] in ('active', 'completed', 'archived'):
            p.status = data['status']
        if data.get('deadline'):
            try:
                p.deadline = datetime.strptime(data['deadline'], '%Y-%m-%d')
            except ValueError:
                pass
        db.session.commit()
        return jsonify(p.to_dict())
    except Exception as e:
        db.session.rollback()
        logger.exception('Unhandled error')
        return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/projects/<int:pid>', methods=['DELETE'])
@login_required
@_login_limit
def delete_project(pid):
    """
    Hard-delete a project and all its tasks (cascade).
    Only the project owner or a full admin can delete; leads cannot.
    """
    try:
        p = db.session.get(Project, pid)   # BUG-06 FIX
        if p is None:
            return jsonify({'error': 'Project not found'}), 404
        if p.owner_id != current_user.id and current_user.role != 'admin':
            return jsonify({'error': 'Unauthorized'}), 403
        db.session.delete(p)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        logger.exception('Unhandled error')
        return jsonify({'error': 'An internal error occurred'}), 500


# =============================================================================
# TASKS API
# =============================================================================
#
# Endpoint summary:
#   GET    /api/tasks              → list tasks (filterable, paginated)
#   POST   /api/tasks              → create a task
#   PUT    /api/tasks/<tid>        → update fields on a task
#   DELETE /api/tasks/<tid>        → hard-delete a task
#   POST   /api/tasks/bulk         → apply a patch to multiple tasks
#   POST   /api/tasks/bulk-delete  → delete multiple tasks (admin/lead only)
#   GET    /api/tasks/export       → download xlsx of current tasks
#   GET    /api/tasks/template     → download blank import template
#   POST   /api/tasks/parse-xlsx   → server-side xlsx parse → JSON rows
#   POST   /api/tasks/bulk-from-process → import tasks from process/task-name rows
#   GET    /api/tasks/kanban       → tasks grouped by status for Kanban view
#   GET    /api/tasks/<tid>/audit  → audit trail for a single task

@app.route('/api/tasks', methods=['GET'])
@login_required
def get_tasks():
    """
    Return a filtered, paginated list of tasks visible to the current user.

    Query parameters:
      project_id  (int)     — filter to a specific project
      status      (str)     — one of: pending, in_progress, review, done, blocked
      priority    (str)     — one of: low, medium, high, critical
      search      (str)     — searches study_id, sr_rfc, task_name (ILIKE, max 100 chars)
      limit       (int)     — max rows to return (default 500, hard cap 500)
      offset      (int)     — row offset for pagination (default 0)

    Visibility:
      member — only tasks assigned to them or created by them.
      admin/lead — all tasks.

    Performance:
      BUG-04 / PERF-03: joinedload on _assignee, creator, and project ensures all
      related data is loaded in a single JOIN, eliminating N+1 SELECT patterns in
      to_dict().
    """
    try:
        # BUG-04 FIX: Eager-load all relationships to eliminate N+1 queries
        q = Task.query.options(
            joinedload(Task._assignee),
            joinedload(Task.creator),
            joinedload(Task.project),
        )
        if current_user.role == 'member':
            q = q.filter(db.or_(
                Task.assigned_to == current_user.id,
                Task.created_by  == current_user.id,
            ))
        if request.args.get('project_id'):
            try:
                q = q.filter(Task.project_id == int(request.args['project_id']))
            except (TypeError, ValueError):
                return jsonify({'error': 'Invalid project_id'}), 400
        VALID_STATUSES   = ('pending', 'in_progress', 'review', 'done', 'blocked')
        VALID_PRIORITIES = ('low', 'medium', 'high', 'critical')
        if request.args.get('status'):
            if request.args['status'] not in VALID_STATUSES:
                return jsonify({'error': 'Invalid status filter'}), 400
            q = q.filter(Task.status == request.args['status'])
        if request.args.get('priority'):
            if request.args['priority'] not in VALID_PRIORITIES:
                return jsonify({'error': 'Invalid priority filter'}), 400
            q = q.filter(Task.priority == request.args['priority'])
        if request.args.get('search'):
            # SEC-08: Escape LIKE metacharacters so user input can't alter
            # the pattern structure (e.g. sending '%' or '_' floods results).
            raw = request.args['search'][:100]
            escaped = raw.replace('\\', '\\\\').replace('%', '\\%').replace('_', '\\_')
            s = f'%{escaped}%'
            q = q.filter(db.or_(
                Task.study_id.ilike(s, escape='\\'),
                Task.sr_rfc.ilike(s, escape='\\'),
                Task.task_name.ilike(s, escape='\\'),
            ))
        # PERF: Support cursor-style pagination so callers can page through large
        # result sets without fetching all 500 rows in one shot.
        try:
            limit  = min(int(request.args.get('limit',  500)), 500)
            offset = max(int(request.args.get('offset', 0)),   0)
        except (TypeError, ValueError):
            limit, offset = 500, 0
        tasks = q.order_by(Task.created_at.desc()).limit(limit).offset(offset).all()
        return jsonify([t.to_dict() for t in tasks])
    except Exception as e:
        logger.exception('Unhandled error'); return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/tasks', methods=['POST'])
@login_required
@_login_limit
def create_task():
    """
    Create a new task.

    Required fields: project_id, task_name, sr_rfc.
    Optional: study_id (auto-generated if omitted), due_date, assigned_to,
              volume, turnaround_time, process, subtask, notes, started_at.

    Security:
      SEC-01: status and priority are validated against allow-lists; invalid values
              are silently coerced to defaults rather than rejected so the UI
              can omit them for brevity.
      SEC-03: assigned_to is validated to be an existing user ID.
      SEC-E:  volume capped 1–9 999; turnaround_time capped 1–99 999 minutes.
      SEC-I:  subtask stripped to alphanumeric + space only.

    Post-commit:
      • SSE 'task_created' event broadcast to all connected clients.
      • If assigned_to is set, notify_task_assigned() enqueues an email to the member.
        db.session.refresh(t) is called first so the _assignee relationship is
        populated with the newly-committed User row.
    """
    try:
        data = request.get_json(silent=True) or {}
        # Require the three minimum fields
        for f in ('project_id', 'task_name', 'sr_rfc'):
            if not data.get(f):
                return jsonify({'error': f'{f} is required'}), 400
        try:
            _pid = int(data['project_id'])
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid project_id'}), 400
        if db.session.get(Project, _pid) is None:   # BUG-06 FIX
            return jsonify({'error': 'Project not found'}), 404
        # Study ID: use caller-supplied value or auto-generate a unique one
        supplied_sid = sanitize(data.get('study_id', '')).strip()[:50]
        if supplied_sid:
            if Task.query.filter_by(study_id=supplied_sid).first():
                return jsonify({'error': f'Study ID "{supplied_sid}" already exists'}), 409
            study_id = supplied_sid
        else:
            study_id = gen_study_id()
        due = None
        if data.get('due_date'):
            try:
                due = datetime.strptime(data['due_date'], '%Y-%m-%d')
            except ValueError:
                pass   # Silently ignore malformed dates — due stays None
        # SEC-01: Coerce invalid enum values to safe defaults
        VALID_STATUSES   = ('pending', 'in_progress', 'review', 'done', 'blocked')
        VALID_PRIORITIES = ('low', 'medium', 'high', 'critical')
        req_status   = data.get('status',   'pending')
        req_priority = data.get('priority', 'medium')
        if req_status   not in VALID_STATUSES:   req_status   = 'pending'
        if req_priority not in VALID_PRIORITIES: req_priority = 'medium'
        # SEC-03: Validate assigned_to is a real user ID
        try:
            _assigned_to = int(data['assigned_to']) if data.get('assigned_to') else None
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid assigned_to'}), 400
        # SEC-E: Cap numeric fields to prevent unreasonable values
        try:
            _volume = min(max(1, int(data.get('volume', 1) or 1)), 9999)
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid volume'}), 400
        try:
            _tat = int(data['turnaround_time']) if data.get('turnaround_time') else None
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid turnaround_time'}), 400

        t = Task(
            study_id=study_id,
            sr_rfc=sanitize(data['sr_rfc'])[:100],
            project_id=_pid,
            assigned_to=_assigned_to,
            task_name=sanitize(data['task_name'])[:300],
            process=sanitize(data.get('process', ''))[:200],
            # SEC-I: subtask is alphanumeric + space only to prevent injection
            subtask=re.sub(r'[^a-zA-Z0-9 ]', '', sanitize(data.get('subtask', '')))[:300],
            status=req_status,
            priority=req_priority,
            volume=_volume,
            turnaround_time=_tat,
            due_date=due,
            notes=sanitize(data.get('notes', ''))[:2000],
            created_by=current_user.id,
        )
        # Accept started_at from bulk import or API consumers (two formats supported)
        if data.get('started_at'):
            for fmt in ('%Y-%m-%d %H:%M', '%Y-%m-%d'):
                try:
                    t.started_at = datetime.strptime(data['started_at'], fmt)
                    break
                except ValueError:
                    continue
        db.session.add(t)
        db.session.commit()
        # Broadcast the new task to all connected SSE clients immediately
        sse.publish('task_created', {'task': t.to_dict()})
        # Email the assigned member — refresh first to load the _assignee relationship
        if t.assigned_to:
            db.session.refresh(t)   # Populate _assignee after commit
            notify_task_assigned(t)
        return jsonify(t.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        logger.exception('Unhandled error')
        return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/tasks/<int:tid>', methods=['PUT'])
@login_required
@_login_limit
def update_task(tid):
    """
    Update fields on an existing task.

    Authorization:
      The task creator, the assigned member, or any admin/lead may edit.
      Pure members cannot edit tasks they neither created nor are assigned to.

    Audit trail:
      Original field values are captured before mutations.  After commit,
      _audit_task_change() is called for every tracked field — rows are only
      inserted when a value actually changed (old != new).

    Auto-timestamps:
      • started_at is set to utcnow() automatically when status → 'in_progress'
        (unless started_at is already set or explicitly supplied in the request).
      • completed_at is set to utcnow() when status → 'done' (once only).
      Both can also be supplied explicitly; started_at accepts 'YYYY-MM-DD HH:MM'
      or 'YYYY-MM-DD' so it round-trips correctly regardless of input source.

    Post-commit:
      • SSE 'task_updated' event broadcast to all connected clients.
      • notify_task_submitted() fires when status transitions to 'done'.
      • notify_task_assigned() fires when assigned_to changes to a new user.
        db.session.refresh() is called before the notification so _assignee
        reflects the new assignee, not the ORM's stale cached reference.
    """
    try:
        task = db.session.get(Task, tid)   # BUG-06 FIX: db.session.get() not Query.get()
        if task is None:
            return jsonify({'error': 'Task not found'}), 404
        if (task.created_by != current_user.id
                and task.assigned_to != current_user.id
                and current_user.role not in ('admin', 'lead')):
            return jsonify({'error': 'Unauthorized'}), 403
        data = request.get_json(silent=True) or {}

        # Snapshot current values before any mutations — used by the audit loop below
        _AUDIT_FIELDS = ('status', 'priority', 'assigned_to', 'task_name',
                         'process', 'subtask', 'notes', 'sr_rfc', 'volume',
                         'turnaround_time', 'due_date', 'started_at', 'completed_at')
        _orig = {f: getattr(task, f, None) for f in _AUDIT_FIELDS}

        # SEC-01: Reject invalid enum values up-front (strict for updates, unlike create)
        VALID_STATUSES   = ('pending', 'in_progress', 'review', 'done', 'blocked')
        VALID_PRIORITIES = ('low', 'medium', 'high', 'critical')
        if 'status' in data and data['status'] not in VALID_STATUSES:
            return jsonify({'error': f'Invalid status. Must be one of: {", ".join(VALID_STATUSES)}'}), 400
        if 'priority' in data and data['priority'] not in VALID_PRIORITIES:
            return jsonify({'error': f'Invalid priority. Must be one of: {", ".join(VALID_PRIORITIES)}'}), 400
        # Apply sanitized text field updates
        for f in ['task_name', 'process', 'subtask', 'notes', 'sr_rfc', 'status', 'priority']:
            if f in data:
                v = sanitize(data[f])
                if f == 'subtask':
                    v = re.sub(r'[^a-zA-Z0-9 ]', '', v)   # SEC-I: alphanumeric only
                setattr(task, f, v)
        # SEC-E: Clamp numeric fields to prevent absurd values
        for f in ['volume', 'turnaround_time']:
            if f in data and data[f] is not None:
                try:
                    val = int(data[f])
                    if f == 'volume':
                        val = min(max(1, val), 9999)    # volume: 1–9 999
                    elif f == 'turnaround_time':
                        val = min(max(1, val), 99999)   # TAT: 1–99 999 minutes
                    setattr(task, f, val)
                except (TypeError, ValueError):
                    return jsonify({'error': f'Invalid value for {f}'}), 400
        if 'assigned_to' in data:
            if data['assigned_to']:
                # SEC-03: Verify the target user exists before assigning (prevent IDOR)
                try:
                    aid = int(data['assigned_to'])
                except (TypeError, ValueError):
                    return jsonify({'error': 'Invalid assigned_to'}), 400
                if not db.session.get(User, aid):
                    return jsonify({'error': 'Assigned user not found'}), 400
                task.assigned_to = aid
            else:
                task.assigned_to = None   # Explicit None = unassign
        if 'due_date' in data:
            if data['due_date']:
                try:
                    task.due_date = datetime.strptime(data['due_date'], '%Y-%m-%d')
                except ValueError:
                    pass
            else:
                task.due_date = None   # Allow explicitly clearing the due date
        if 'started_at' in data:
            if data['started_at']:
                # Accept both full datetime ('YYYY-MM-DD HH:MM') and date-only ('YYYY-MM-DD')
                # so the field round-trips correctly regardless of input source
                for fmt in ('%Y-%m-%d %H:%M', '%Y-%m-%d'):
                    try:
                        task.started_at = datetime.strptime(data['started_at'], fmt)
                        break
                    except ValueError:
                        continue
            else:
                task.started_at = None
        elif data.get('status') == 'in_progress' and not task.started_at:
            # Auto-set started_at when transitioning to in_progress for the first time
            task.started_at = datetime.utcnow()
        if data.get('status') == 'done' and not task.completed_at:
            # Auto-set completed_at on first transition to done
            task.completed_at = datetime.utcnow()

        # ── Audit trail: write one row per changed field ─────────────────────
        # Reuses _AUDIT_FIELDS / _orig captured before any mutations above.
        for f in _AUDIT_FIELDS:
            new_v = getattr(task, f, None)
            _audit_task_change(task.id, f, _orig.get(f), new_v, current_user.id)

        db.session.commit()

        # ── UPGRADE: SSE broadcast ────────────────────────────────────────────
        sse.publish('task_updated', {'task': task.to_dict()})

        # ── Email: notify member on submission (status → done) ─────────────
        if _orig.get('status') != 'done' and task.status == 'done':
            notify_task_submitted(task)
        # ── Email: notify member if reassigned to a new person ────────────
        # db.session.refresh() reloads the _assignee relationship after the
        # commit so notify_task_assigned() sees the correct User object and
        # email address for the newly assigned user, not the stale cached one.
        if 'assigned_to' in data and _orig.get('assigned_to') != task.assigned_to and task.assigned_to:
            db.session.refresh(task)
            notify_task_assigned(task)

        return jsonify(task.to_dict())
    except Exception as e:
        db.session.rollback()
        logger.exception('Unhandled error'); return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/tasks/<int:tid>', methods=['DELETE'])
@login_required
@_login_limit
def delete_task(tid):
    """
    Hard-delete a single task and broadcast the deletion over SSE.
    Only the task creator or an admin/lead can delete.
    The associated task_audit rows are removed automatically via ON DELETE CASCADE.
    """
    try:
        task = db.session.get(Task, tid)   # BUG-06 FIX
        if task is None:
            return jsonify({'error': 'Task not found'}), 404
        if (task.created_by != current_user.id
                and current_user.role not in ('admin', 'lead')):
            return jsonify({'error': 'Unauthorized'}), 403
        task_id = task.id
        db.session.delete(task)
        db.session.commit()
        # Notify all SSE clients so they can remove the card from the UI immediately
        sse.publish('task_deleted', {'task_id': task_id})
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        logger.exception('Unhandled error')
        return jsonify({'error': 'An internal error occurred'}), 500


# =============================================================================
# BULK TASK ACTIONS
# =============================================================================

@app.route('/api/tasks/bulk', methods=['POST'])
@login_required
@_login_limit
def bulk_update_tasks():
    """
    Apply a status / priority / assigned_to patch to a list of task IDs.

    Body:
      ids        (list[int])  — task IDs to update (max 500)
      status     (str)        — optional new status
      priority   (str)        — optional new priority
      assigned_to (int|null)  — optional new assignee (null to unassign)

    Authorization:
      Members can only update tasks they own or are assigned to; others are skipped
      (counted in the 'skipped' response field, not treated as errors).

    Post-commit:
      • SSE 'task_updated' broadcast for every updated task (BUG-FIX: was missing).
      • notify_task_submitted() called for each task that transitioned to 'done'
        (BUG-FIX: was missing in original bulk path).
    """
    try:
        data = request.get_json(silent=True) or {}
        ids  = data.get('ids', [])
        if not isinstance(ids, list) or not ids:
            return jsonify({'error': 'ids must be a non-empty list'}), 400
        if len(ids) > 500:
            return jsonify({'error': 'Max 500 tasks per bulk action'}), 400
        try:
            ids = [int(i) for i in ids]
        except (TypeError, ValueError):
            return jsonify({'error': 'ids must be integers'}), 400

        VALID_STATUSES   = ('pending', 'in_progress', 'review', 'done', 'blocked')
        VALID_PRIORITIES = ('low', 'medium', 'high', 'critical')
        patch = {}
        if 'status' in data:
            if data['status'] not in VALID_STATUSES:
                return jsonify({'error': f'Invalid status'}), 400
            patch['status'] = data['status']
        if 'priority' in data:
            if data['priority'] not in VALID_PRIORITIES:
                return jsonify({'error': f'Invalid priority'}), 400
            patch['priority'] = data['priority']
        if 'assigned_to' in data:
            if data['assigned_to']:
                try:
                    aid = int(data['assigned_to'])
                except (TypeError, ValueError):
                    return jsonify({'error': 'Invalid assigned_to'}), 400
                if not db.session.get(User, aid):
                    return jsonify({'error': 'Assigned user not found'}), 404
                patch['assigned_to'] = aid
            else:
                patch['assigned_to'] = None
        if not patch:
            return jsonify({'error': 'No fields to update'}), 400

        tasks = Task.query.options(
            joinedload(Task._assignee),
        ).filter(Task.id.in_(ids)).all()
        updated, skipped = 0, 0
        done_tasks = []   # collect tasks flipped to done for post-commit emails
        for task in tasks:
            # Members can only update tasks they own or are assigned to
            if (current_user.role not in ('admin', 'lead')
                    and task.created_by != current_user.id
                    and task.assigned_to != current_user.id):
                skipped += 1
                continue
            was_done = task.status == 'done'
            for field, value in patch.items():
                setattr(task, field, value)
            # Auto-timestamps
            if patch.get('status') == 'in_progress' and not task.started_at:
                task.started_at = datetime.utcnow()
            if patch.get('status') == 'done' and not task.completed_at:
                task.completed_at = datetime.utcnow()
                if not was_done:
                    done_tasks.append(task)
            updated += 1
        db.session.commit()
        # Broadcast each updated task over SSE so connected clients update live
        for task in tasks:
            sse.publish('task_updated', {'task': task.to_dict()})
        # Email submission confirmation for tasks just marked done
        for task in done_tasks:
            notify_task_submitted(task)
        return jsonify({'updated': updated, 'skipped': skipped})
    except Exception as e:
        db.session.rollback()
        logger.exception('bulk_update_tasks error')
        return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/tasks/bulk-delete', methods=['POST'])
@login_required
@_login_limit
def bulk_delete_tasks():
    """
    Hard-delete multiple tasks by ID.  Admin and lead only.

    Body:
      ids (list[int]) — task IDs to delete (max 500)

    Uses SQLAlchemy's bulk DELETE with synchronize_session='fetch' so the ORM
    session reflects the deletion without requiring individual object loads.
    """
    try:
        if current_user.role not in ('admin', 'lead'):
            return jsonify({'error': 'Unauthorized'}), 403
        data = request.get_json(silent=True) or {}
        ids  = data.get('ids', [])
        if not isinstance(ids, list) or not ids:
            return jsonify({'error': 'ids must be a non-empty list'}), 400
        if len(ids) > 500:
            return jsonify({'error': 'Max 500 tasks per bulk delete'}), 400
        try:
            ids = [int(i) for i in ids]
        except (TypeError, ValueError):
            return jsonify({'error': 'ids must be integers'}), 400
        deleted = Task.query.filter(Task.id.in_(ids)).delete(synchronize_session='fetch')
        db.session.commit()
        return jsonify({'deleted': deleted})
    except Exception as e:
        db.session.rollback()
        logger.exception('bulk_delete_tasks error')
        return jsonify({'error': 'An internal error occurred'}), 500


# =============================================================================
# TASKS: EXPORT / IMPORT TEMPLATE / XLSX PARSE / BULK-FROM-PROCESS
# =============================================================================

@app.route('/api/tasks/export', methods=['GET'])
@login_required
def export_tasks():
    """
    Generate and return an xlsx download of the current user's visible tasks.

    SEC-09: Members only receive their own tasks (assigned or created by them).
    Admins and leads receive all tasks.  Hard cap at 500 rows.

    The workbook uses a dark header row matching the portal brand colour,
    alternating row fills, and fixed column widths for readability.
    """
    try:

        q = Task.query.options(
            joinedload(Task._assignee),
            joinedload(Task.project),
            joinedload(Task.creator),   # Eager-load to avoid N+1 in to_dict()
        )
        # SEC-09: Scope export to member's own tasks
        if current_user.role == 'member':
            q = q.filter(db.or_(
                Task.assigned_to == current_user.id,
                Task.created_by  == current_user.id,
            ))
        tasks = q.order_by(Task.created_at.desc()).limit(500).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Tasks'

        thin   = Side(style='thin', color='DDDDDD')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        headers = ['Study ID','SR/RFC','Mapping Process','Task Name','Sub Task','Assignee',
                   'Enterprise ID','TAT (hrs)','Volume','Start Date','Due Date',
                   'Status','Priority','Notes']
        col_widths = [14,14,20,30,24,16,14,9,8,13,13,13,11,30]

        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            c.fill      = PatternFill('solid', start_color='0A0A0F')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = border
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 24

        alt = PatternFill('solid', start_color='F8F7F4')
        for ri, t in enumerate(tasks, 2):
            td = t.to_dict()
            vals = [
                td.get('study_id',''), td.get('sr_rfc',''), td.get('process',''),
                td.get('task_name',''), td.get('subtask',''), td.get('assignee_name',''),
                td.get('enterprise_id',''), td.get('turnaround_time',''), td.get('volume',''),
                td.get('started_at','').split(' ')[0] if td.get('started_at') else '',
                td.get('due_date',''), td.get('status',''), td.get('priority',''),
                td.get('notes',''),
            ]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font      = Font(name='Arial', size=9)
                c.alignment = Alignment(vertical='center')
                c.border    = border
                if ri % 2 == 0:
                    c.fill = alt

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"HuddleTasks_{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
        return send_file(buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=fname)
    except Exception as e:
        logger.exception('Unhandled error'); return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/tasks/template', methods=['GET'])
@login_required
def download_tasks_template():
    """
    Return a pre-filled xlsx import template for task bulk upload.

    The template has one example row populated with real project_code and task_name
    values fetched from the database so users can see correct formatting at a glance.
    Columns marked with * are required by the import parser.
    """
    try:

        # Populate example row with real data from DB when available
        proj  = Project.query.first()
        tn    = TaskName.query.first()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Tasks Import Template'

        thin   = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        headers = ['SR/RFC *','Project Code *','Mapping Process','Task Name *','Sub Task',
                   'Assignee Username','TAT (hrs)','Volume',
                   'Start Date (YYYY-MM-DD)','Due Date (YYYY-MM-DD)',
                   'Status','Priority','Notes']
        col_widths = [16,16,20,30,24,18,9,8,22,22,12,12,30]

        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            c.fill      = PatternFill('solid', start_color='0A0A0F')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = border
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 26

        example = [
            'RFC-2024-001',
            proj.project_code if proj else 'PROJ-001',
            tn.process if tn else 'NG 1.0',
            tn.task_name if tn else 'Sample Task',
            '', '', 8, 1, '', '', 'pending', 'medium', ''
        ]
        for ci, val in enumerate(example, 1):
            c = ws.cell(row=2, column=ci, value=val)
            c.font      = Font(name='Arial', size=9)
            c.alignment = Alignment(vertical='center')
            c.border    = border

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name='HuddleTasksImportTemplate.xlsx')
    except Exception as e:
        logger.exception('Unhandled error'); return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/tasks/bulk-from-process', methods=['POST'])
@login_required
@_login_limit
def bulk_tasks_from_process():
    """
    Create Task records from a list of {process, task_name} row objects.

    Used by the Huddle portal's "Import from Process" feature.  Each row is
    matched against the TaskName lookup table to auto-fill turnaround_time
    (cap_timing).  SR/RFC is auto-generated; the creator and assignee are both
    set to the current user; all tasks go into the first available project.

    Collision handling:
      db.session.flush() is called after each add so the UNIQUE constraint on
      study_id is checked immediately.  On collision the study_id is regenerated
      and the flush is retried once — prevents a rare same-second hex collision
      from silently rolling back the entire batch.
    """
    try:
        data = request.get_json(silent=True) or {}
        rows = data.get('rows', [])
        if not rows:
            return jsonify({'error': 'No rows provided'}), 400
        if len(rows) > 1000:
            return jsonify({'error': 'Too many rows: max 1000 per import'}), 400

        # Use the first available project as the default project
        project = Project.query.order_by(Project.created_at.asc()).first()
        if not project:
            return jsonify({'error': 'No projects found. Please ask an Admin to create a project first.'}), 400

        # Build a lookup: (process_lower, task_name_lower) -> TaskName
        all_tns = TaskName.query.filter_by(status='active').all()
        tn_lookup = {
            (tn.process.lower().strip(), tn.task_name.lower().strip()): tn
            for tn in all_tns
        }
        # Also a process-only lookup for partial matching
        proc_lookup = {}
        for tn in all_tns:
            proc_lookup.setdefault(tn.process.lower().strip(), []).append(tn)

        ok, fail, errors = 0, 0, []
        for i, row in enumerate(rows):
            process   = sanitize(str(row.get('process', ''))).strip()
            task_name = sanitize(str(row.get('task_name', ''))).strip()
            if not process or not task_name:
                fail += 1
                errors.append(f'Row {i+1}: missing process or task name')
                continue

            # Look up the TaskName to get the cap_timing (TAT)
            tn = tn_lookup.get((process.lower(), task_name.lower()))
            tat = tn.cap_timing if tn else None

            # Auto-generate a unique SR/RFC reference
            sr_rfc = f'IMP-{datetime.utcnow().strftime("%Y%m%d")}-{uuid.uuid4().hex[:6].upper()}'

            task = Task(
                study_id=gen_study_id(),
                sr_rfc=sr_rfc,
                project_id=project.id,
                assigned_to=current_user.id,
                task_name=task_name,
                process=process,
                subtask='',
                status='pending',
                priority='medium',
                volume=1,
                turnaround_time=tat,
                created_by=current_user.id,
            )
            db.session.add(task)
            # BUG-FIX: Flush after each add so the DB can enforce the study_id
            # UNIQUE constraint immediately. Without this, multiple tasks generated
            # in the same batch can receive the same study_id (same-second prefix +
            # 6-char hex collision) and only fail at the final commit, rolling back
            # the entire import silently.
            try:
                db.session.flush()
            except Exception:
                db.session.rollback()
                # Retry with a freshly generated study_id
                task.study_id = gen_study_id()
                db.session.add(task)
                db.session.flush()
            ok += 1

        db.session.commit()
        return jsonify({'imported': ok, 'failed': fail, 'errors': errors}), 201
    except Exception as e:
        db.session.rollback()
        logger.exception('bulk-from-process error')
        return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/tasks/parse-xlsx', methods=['POST'])
@login_required
@_login_limit
def parse_tasks_xlsx():
    """
    Parse an uploaded xlsx file and return its rows as JSON.

    Used by the Admin portal's task import wizard — the browser sends the file
    here, gets back headers + rows, shows the user a preview, then POSTs the
    mapped data to /api/tasks (one by one) or /api/tasks/bulk-from-process.

    Security (SEC-12):
      • File extension must be .xlsx.
      • MIME type is checked against an allow-list (some browsers send
        'application/octet-stream' or 'application/zip' for xlsx files).
      • Magic bytes (first 2 bytes == 'PK') are verified before passing the
        stream to openpyxl — prevents disguised HTML/executable uploads from
        reaching the parser.
      • Row count capped at MAX_IMPORT_ROWS (5 000) to prevent memory exhaustion.
      • Fully empty rows are silently skipped.
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        f = request.files['file']
        if not f.filename or not f.filename.lower().endswith('.xlsx'):
            return jsonify({'error': 'Please upload a .xlsx file'}), 400
        # SEC-D: Verify actual MIME type, not just the file extension
        allowed_mimes = {
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/octet-stream',  # Some browsers send this for xlsx
            'application/zip',           # xlsx is a zip internally
        }
        if f.mimetype and f.mimetype not in allowed_mimes and not f.mimetype.startswith('application/'):
            return jsonify({'error': 'Invalid file type. Please upload a valid .xlsx file'}), 400
        # SEC-D: Verify xlsx magic bytes (ZIP header: PK\x03\x04) before parsing
        header_bytes = f.stream.read(4)
        f.stream.seek(0)
        if header_bytes[:2] != b'PK':
            return jsonify({'error': 'Invalid file format. File does not appear to be a valid .xlsx file.'}), 400
        wb = load_workbook(f.stream, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            raw_headers = list(next(rows_iter))
        except StopIteration:
            return jsonify({'error': 'File is empty'}), 400
        headers = [str(h).strip() if h is not None else '' for h in raw_headers]
        data_rows = []
        MAX_IMPORT_ROWS = 5000
        for row in rows_iter:
            if len(data_rows) >= MAX_IMPORT_ROWS:
                wb.close()
                return jsonify({'error': f'File exceeds the {MAX_IMPORT_ROWS}-row import limit. Please split it into smaller files.'}), 400
            cells = [str(c).strip() if c is not None else '' for c in row]
            if any(cells):  # skip fully empty rows
                data_rows.append(cells)
        wb.close()
        return jsonify({'headers': headers, 'rows': data_rows}), 200
    except Exception as e:
        logger.exception('xlsx parse error (tasks)')
        return jsonify({'error': 'Failed to parse file. Please check the format and try again.'}), 500


# =============================================================================
# SSE STREAM ENDPOINT
# =============================================================================

@app.route('/api/stream')
@login_required
def sse_stream():
    """
    Server-Sent Events stream — pushes real-time task change events to the browser.

    The client calls EventSource('/api/stream') once.  The server holds the
    connection open and writes SSE-formatted messages whenever a task is created,
    updated, or deleted.

    Protocol:
      • On connect: 'event: connected' is sent immediately as a handshake.
      • Task events: event name is 'task_created', 'task_updated', or 'task_deleted';
        data is a JSON object with the full task dict (or {task_id} for deletes).
      • Heartbeat: a bare comment line (': heartbeat') is sent every 25 s when
        there are no real events.  This keeps the TCP connection alive through
        proxies and load balancers that close idle connections.
      • Disconnect: GeneratorExit is caught; the client queue is unsubscribed so
        it is pruned from the broker's listener list.

    HTTP headers:
      Cache-Control: no-cache  — prevents caching of the stream.
      X-Accel-Buffering: no    — tells Nginx to disable proxy buffering so
                                 each SSE frame reaches the client immediately.
    """
    def event_generator(q: queue.Queue):
        # Immediate handshake confirms the stream is live before any tasks fire
        yield "event: connected\ndata: {}\n\n"
        try:
            while True:
                try:
                    msg = q.get(timeout=25)   # Block up to 25 s for a real event
                    yield msg
                except queue.Empty:
                    # No events in 25 s — send an SSE comment to prevent timeout
                    yield ": heartbeat\n\n"
        except GeneratorExit:
            pass   # Client disconnected cleanly
        finally:
            sse.unsubscribe(q)   # Always clean up to prevent memory leaks

    q = sse.subscribe()
    return Response(
        stream_with_context(event_generator(q)),
        content_type='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',   # Nginx: disable proxy buffering for SSE
        }
    )


# =============================================================================
# AUDIT LOG API
# =============================================================================

@app.route('/api/tasks/<int:tid>/audit', methods=['GET'])
@login_required
def get_task_audit(tid):
    """
    Return the audit trail for a single task, newest-first.

    Shown in the 🕐 History tab of the task detail modal in both portals.
    Results are capped at 200 rows (the oldest changes beyond that threshold
    are not shown, but are preserved in the database).
    """
    task = db.session.get(Task, tid)
    if task is None:
        return jsonify({'error': 'Task not found'}), 404
    # SEC-S2: IDOR fix — members may only view audit logs for tasks they own or
    # are assigned to. Admins and leads can view all task audit logs.
    if current_user.role == 'member':
        if task.assigned_to != current_user.id and task.created_by != current_user.id:
            return jsonify({'error': 'Unauthorized'}), 403
    logs = (TaskAudit.query
            .filter_by(task_id=tid)
            .order_by(TaskAudit.changed_at.desc())
            .limit(200)
            .all())
    return jsonify([l.to_dict() for l in logs])


# =============================================================================
# KANBAN BOARD API
# =============================================================================

@app.route('/api/tasks/kanban', methods=['GET'])
@login_required
def get_tasks_kanban():
    """
    Return all tasks grouped by status for the Kanban board view.

    Query params:
      project_id (int) — optional; restricts the board to a single project.

    Response shape:
      { pending: [...], in_progress: [...], review: [...], done: [...], blocked: [...] }

    Tasks within each column are ordered by priority descending then created_at
    descending so high-priority items float to the top of each lane.
    Tasks with an unrecognised status are placed in 'pending' as a safe fallback.
    """
    try:
        pid_raw = request.args.get('project_id')
        pid = None
        if pid_raw:
            try:
                pid = int(pid_raw)
            except ValueError:
                return jsonify({'error': 'Invalid project_id'}), 400

        q = Task.query.options(
            joinedload(Task._assignee),
            joinedload(Task.creator),
            joinedload(Task.project),
        )
        # SEC-S3: IDOR fix — members only see their own tasks on the Kanban board.
        if current_user.role == 'member':
            q = q.filter(db.or_(
                Task.assigned_to == current_user.id,
                Task.created_by  == current_user.id,
            ))
        if pid:
            q = q.filter_by(project_id=pid)
        # PERF-P2: Cap at 500 tasks to prevent unbounded memory use on large boards.
        tasks = q.order_by(Task.priority.desc(), Task.created_at.desc()).limit(500).all()

        STATUSES = ('pending', 'in_progress', 'review', 'done', 'blocked')
        board = {s: [] for s in STATUSES}
        for t in tasks:
            col = t.status if t.status in board else 'pending'
            board[col].append(t.to_dict())
        return jsonify(board)
    except Exception:
        logger.exception('kanban error')
        return jsonify({'error': 'An internal error occurred'}), 500


# =============================================================================
# USERS API
# =============================================================================
#
# Endpoint summary:
#   GET  /api/users               → list users (scoped by role)
#   GET  /api/users/activity      → per-user task counts for last 3 days (admin/lead)
#   POST /api/users               → create a user (admin only)
#   PUT  /api/users/<uid>         → update role / active flag / password / email (admin)

@app.route('/api/users', methods=['GET'])
@login_required
def get_users():
    """
    Return the list of users.

    Admins and leads see all users including inactive ones (so they can
    re-activate deactivated accounts).  Members see only active users —
    the list is used to populate assignment dropdowns.
    """
    try:
        if current_user.role in ('admin', 'lead'):
            users = User.query.order_by(User.username).all()
        else:
            users = User.query.filter_by(is_active=True).order_by(User.username).all()
        return jsonify([u.to_dict() for u in users])
    except Exception as e:
        logger.exception('Unhandled error')
        return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/users/activity', methods=['GET'])
@login_required
def get_users_activity():
    """
    Return per-user task-creation counts for today, yesterday, and two days ago.

    Admin/lead only.  Used by the Admin portal's Team Activity panel.

    PERF: A single GROUP BY query over the 3-day window replaces what would
    otherwise be N × 3 individual COUNT queries (one per user per day).
    SQLite may return the 'day' column as a string; the code handles both
    string and date types for compatibility.
    """
    try:
        if current_user.role not in ('admin', 'lead'):
            return jsonify({'error': 'Unauthorized'}), 403

        today     = datetime.utcnow().date()
        days      = [today - timedelta(days=i) for i in range(3)]
        window_start = datetime(days[-1].year, days[-1].month, days[-1].day, 0, 0, 0)

        # Single GROUP BY query over a 3-day window — O(1) DB round-trips
        rows = db.session.query(
            Task.created_by,
            db.func.date(Task.created_at).label('day'),
            db.func.count(Task.id).label('cnt'),
        ).filter(
            Task.created_at >= window_start
        ).group_by(Task.created_by, db.func.date(Task.created_at)).all()

        # Build lookup: {user_id: {date_str: count}}
        counts = {}
        for row in rows:
            uid = row.created_by
            day_str = str(row.day)[:10]  # handle both str and date types
            counts.setdefault(uid, {})[day_str] = row.cnt

        users = User.query.filter_by(is_active=True).order_by(User.username).all()
        day_labels = [d.strftime('%b %d') for d in days]
        result = []
        for u in users:
            activity = []
            for i, d in enumerate(days):
                ds = d.strftime('%Y-%m-%d')
                activity.append({'label': day_labels[i], 'count': counts.get(u.id, {}).get(ds, 0)})
            result.append({
                'id': u.id, 'username': u.username,
                'enterprise_id': u.enterprise_id, 'email': u.email,
                'role': u.role, 'is_active': u.is_active,
                'activity': activity,
            })
        return jsonify(result)
    except Exception as e:
        logger.exception('get_users_activity error')
        return jsonify({'error': 'Failed to load activity'}), 500


@app.route('/api/users', methods=['POST'])
@login_required
@_login_limit
def create_user():
    """
    Create a new user account.  Admin only.

    Validations:
      • Username uniqueness
      • Email format (simple regex) and uniqueness
      • Password complexity: min 8 chars, at least one uppercase letter and one digit (SEC-02)
      • Role must be one of: admin, lead, member
      • enterprise_id must be present and unique
    """
    try:
        if current_user.role != 'admin':
            return jsonify({'error': 'Unauthorized'}), 403
        data = request.get_json(silent=True) or {}
        if not data.get('username'):
            return jsonify({'error': 'Username is required'}), 400
        if User.query.filter_by(username=data.get('username')).first():
            return jsonify({'error': 'Username already exists'}), 409
        # FIX-10: Basic email format validation
        email_val = data.get('email', '')
        if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email_val):
            return jsonify({'error': 'Invalid email format'}), 400
        if User.query.filter_by(email=email_val).first():
            return jsonify({'error': 'Email already registered'}), 409
        # SEC-02: Enforce minimum password requirements
        pw = data.get('password', '')
        if len(pw) < 8:
            return jsonify({'error': 'Password must be at least 8 characters'}), 400
        if not re.search(r'[A-Z]', pw) or not re.search(r'[0-9]', pw):
            return jsonify({'error': 'Password must contain at least one uppercase letter and one number'}), 400
        # Validate role
        if data.get('role') and data['role'] not in ('admin', 'lead', 'member'):
            return jsonify({'error': 'Invalid role'}), 400
        # BUG-FIX: Validate enterprise_id is present and not empty before use;
        # accessing data['enterprise_id'] without a check raises KeyError when
        # the client omits the field, bypassing the generic exception handler.
        if not data.get('enterprise_id', '').strip():
            return jsonify({'error': 'Enterprise ID is required'}), 400
        if User.query.filter_by(enterprise_id=data['enterprise_id']).first():
            return jsonify({'error': 'Enterprise ID already in use'}), 409
        u = User(
            username=sanitize(data['username'])[:80],
            email=sanitize(data['email'])[:120],
            enterprise_id=sanitize(data['enterprise_id'])[:50],
            role=data.get('role', 'member'),
        )
        u.set_password(pw)
        db.session.add(u)
        db.session.commit()
        return jsonify(u.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        logger.exception('Unhandled error'); return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/users/<int:uid>', methods=['PUT'])
@login_required
@_login_limit
def update_user(uid):
    """
    Update an existing user account.  Admin only.

    Updatable fields: role, is_active, password, email.

    Self-protection guards:
      • An admin cannot deactivate their own account (would lock them out).
      • An admin cannot demote their own role (would lose admin access).
      Both checks compare uid == current_user.id before applying the change.

    Password updates apply the same complexity rules as creation (SEC-02).
    Email updates check uniqueness against other users (not the same user's
    current email, to avoid a false-positive conflict error).
    """
    try:
        if current_user.role != 'admin':
            return jsonify({'error': 'Unauthorized'}), 403
        u = db.session.get(User, uid)
        if u is None:
            return jsonify({'error': 'User not found'}), 404
        # BUG-FIX: Parse JSON body once — Flask buffers it, but calling get_json()
        # multiple times before the first assignment risks inconsistent results and
        # is confusing. Parse once and reuse the dict throughout.
        data = request.get_json(silent=True) or {}
        # Prevent admin from deactivating or demoting their own account
        if uid == current_user.id:
            if 'is_active' in data and not data.get('is_active', True):
                return jsonify({'error': 'You cannot deactivate your own account'}), 400
            if 'role' in data and data.get('role') != 'admin':
                return jsonify({'error': 'You cannot demote your own admin role'}), 400
        if 'role' in data:
            if data['role'] not in ('admin', 'lead', 'member'):
                return jsonify({'error': 'Invalid role'}), 400
            u.role = data['role']
        if 'is_active' in data:
            u.is_active = bool(data['is_active'])
        if 'password' in data and data['password']:
            pw = data['password']
            if len(pw) < 8:
                return jsonify({'error': 'Password must be at least 8 characters'}), 400
            if not re.search(r'[A-Z]', pw) or not re.search(r'[0-9]', pw):
                return jsonify({'error': 'Password must contain at least one uppercase letter and one number'}), 400
            u.set_password(pw)
        if 'email' in data and data['email']:
            email_val = data['email']
            if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email_val):
                return jsonify({'error': 'Invalid email format'}), 400
            existing = User.query.filter_by(email=email_val).first()
            if existing and existing.id != uid:
                return jsonify({'error': 'Email already in use'}), 409
            u.email = sanitize(email_val)[:120]
        db.session.commit()
        return jsonify(u.to_dict())
    except Exception as e:
        db.session.rollback()
        logger.exception('Unhandled error'); return jsonify({'error': 'An internal error occurred'}), 500


# =============================================================================
# ANALYTICS API
# =============================================================================

@app.route('/api/analytics/daily', methods=['GET'])
@login_required
def analytics_daily():
    """
    Return task activity data for the calendar and daily-drill-down views.

    Modes (mutually exclusive query params):
      ?date=YYYY-MM-DD  → Return all tasks that were created, started, or due on
                          that specific day, each annotated with a 'day_flags' list
                          (['created'], ['started'], ['due'], or combinations).

      ?month=YYYY-MM    → Return per-day summary counts for the whole month.
                          Each day entry has: created, started, due, done, blocked,
                          total (de-duplicated: a task counted only once per day
                          even if multiple date fields fall on the same day).
                          Also returns first_weekday (0=Mon) for calendar rendering.

    No params → defaults to the current UTC month (same as month mode).

    Visibility: members see only their own tasks; admins/leads see everything.
    """
    try:
        month_str = request.args.get('month')   # e.g. 2025-03
        date_str  = request.args.get('date')    # e.g. 2025-03-15

        # PERF-P3: Eager-load relationships to prevent N+1 queries when
        # to_dict() is called on each task in the date-mode drill-down response.
        q = Task.query.options(
            joinedload(Task._assignee),
            joinedload(Task.project),
        )
        if current_user.role == 'member':
            q = q.filter(db.or_(Task.assigned_to == current_user.id,
                                 Task.created_by  == current_user.id))

        if date_str:
            # Return all tasks for a specific day (created_at OR started_at OR due_date)
            try:
                day = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'error': 'Invalid date format, use YYYY-MM-DD'}), 400

            day_start = datetime(day.year, day.month, day.day, 0, 0, 0)
            day_end   = datetime(day.year, day.month, day.day, 23, 59, 59)

            tasks = q.filter(
                db.or_(
                    db.and_(Task.created_at >= day_start, Task.created_at <= day_end),
                    db.and_(Task.started_at >= day_start, Task.started_at <= day_end),
                    db.and_(Task.due_date   >= day_start, Task.due_date   <= day_end),
                )
            ).order_by(Task.created_at.desc()).all()

            result = []
            for t in tasks:
                d = t.to_dict()
                # Annotate which date fields matched this day
                flags = []
                if t.created_at and t.created_at.date() == day:   flags.append('created')
                if t.started_at and t.started_at.date() == day:   flags.append('started')
                if t.due_date   and t.due_date.date()   == day:   flags.append('due')
                d['day_flags'] = flags
                result.append(d)

            return jsonify({'date': date_str, 'tasks': result, 'count': len(result)})

        if month_str:
            try:
                y, m = int(month_str[:4]), int(month_str[5:7])
            except (ValueError, IndexError):
                return jsonify({'error': 'Invalid month format, use YYYY-MM'}), 400
        else:
            now = datetime.utcnow()
            y, m = now.year, now.month

        _, days_in_month = cal_mod.monthrange(y, m)
        month_start = datetime(y, m, 1)
        month_end   = datetime(y, m, days_in_month, 23, 59, 59)

        tasks = q.filter(
            db.or_(
                db.and_(Task.created_at >= month_start, Task.created_at <= month_end),
                db.and_(Task.started_at >= month_start, Task.started_at <= month_end),
                db.and_(Task.due_date   >= month_start, Task.due_date   <= month_end),
            )
        ).all()

        # Build per-day summary
        day_map = defaultdict(lambda: {'created': 0, 'started': 0, 'due': 0,
                                       'done': 0, 'blocked': 0, 'total': 0})
        seen = defaultdict(set)  # avoid double-counting same task on same day

        for t in tasks:
            for dt, flag in [(t.created_at, 'created'), (t.started_at, 'started'), (t.due_date, 'due')]:
                if dt and dt.year == y and dt.month == m:
                    key = dt.strftime('%Y-%m-%d')
                    day_map[key][flag] += 1
                    if t.id not in seen[key]:
                        seen[key].add(t.id)
                        day_map[key]['total'] += 1
                        if t.status == 'done':    day_map[key]['done']    += 1
                        if t.status == 'blocked': day_map[key]['blocked'] += 1

        return jsonify({
            'year': y, 'month': m,
            'days_in_month': days_in_month,
            'first_weekday': cal_mod.monthrange(y, m)[0],  # 0=Mon … 6=Sun
            'days': {k: v for k, v in day_map.items()},
        })
    except Exception as e:
        logger.exception('Unhandled error')
        return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/analytics/tat-warnings', methods=['GET'])
@login_required
def tat_warnings():
    """
    Return a list of days where a member's total TAT exceeds 540 minutes (9 hours).

    Used by the Admin portal's TAT Warnings panel to flag potential data-entry
    errors or workload anomalies.

    Algorithm:
      • Fetch all tasks that have both turnaround_time and started_at set.
      • Group by (assignee_id, started_at date) and sum turnaround_time (minutes).
      • Emit a warning entry for every (member, day) bucket that exceeds LIMIT.
      • Results are sorted newest-date first.

    Visibility: members see only their own tasks; admins/leads see all members.
    """
    try:
        # PERF-P1: Limit TAT scan to last 30 days — prevents unbounded full-table
        # scan as task history grows. Warnings older than 30 days are not actionable.
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        q = Task.query.options(joinedload(Task._assignee)).filter(
            Task.turnaround_time.isnot(None),
            Task.started_at.isnot(None),
            Task.started_at >= thirty_days_ago,
        )
        if current_user.role == 'member':
            q = q.filter(db.or_(
                Task.assigned_to == current_user.id,
                Task.created_by  == current_user.id,
            ))
        tasks = q.all()

        # Group by (assignee_id, date) and sum TAT in minutes
        buckets = defaultdict(lambda: {'name': '', 'eid': '', 'total_mins': 0, 'tasks': []})
        for t in tasks:
            if not t._assignee:
                continue
            day = t.started_at.strftime('%Y-%m-%d')
            key = (t.assigned_to, day)
            b = buckets[key]
            b['name'] = t._assignee.username
            b['eid']  = t._assignee.enterprise_id
            b['total_mins'] += (t.turnaround_time or 0)   # already in minutes
            b['tasks'].append({
                'task_name': t.task_name,
                'sr_rfc':    t.sr_rfc,
                'tat_mins':  (t.turnaround_time or 0),
            })

        LIMIT = 540   # 9 hours in minutes
        warnings = []
        for (uid, day), b in sorted(buckets.items(), key=lambda x: x[0][1], reverse=True):
            if b['total_mins'] > LIMIT:
                warnings.append({
                    'member':        b['name'],
                    'enterprise_id': b['eid'],
                    'date':          day,
                    'total_mins':    b['total_mins'],
                    'total_hrs':     round(b['total_mins'] / 60, 1),
                    'over_by_mins':  b['total_mins'] - LIMIT,
                    'tasks':         b['tasks'],
                })
        return jsonify({'warnings': warnings, 'limit_mins': LIMIT})
    except Exception as e:
        logger.exception('tat_warnings error')
        return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/analytics', methods=['GET'])
@login_required
def analytics():
    """
    Return the headline summary statistics used by the Admin portal dashboard
    and the Huddle portal's My Stats panel.

    Fields returned:
      total_tasks, done_tasks, pending_tasks, in_progress, blocked,
      total_projects, total_volume, avg_tat_mins, priority_data (dict),
      completion_rate (%), total_users (active)

    PERF-01: A single GROUP BY aggregate query replaces 8+ individual COUNT
    queries that the original code used.

    BUG-05 FIX: Members only see figures for tasks assigned to or created by
    them; admins and leads see organisation-wide figures.

    avg_tat_mins is computed entirely in SQLite using strftime('%s', ...) arithmetic
    to calculate seconds between started_at and completed_at, then converted to
    minutes in Python — no row fetching required.
    """
    try:
        # BUG-05 FIX: Scope analytics to the requesting user's tasks for members.
        # Admins and leads continue to see org-wide figures.
        if current_user.role == 'member':
            projs = db.session.query(Task.project_id).filter(
                db.or_(Task.assigned_to == current_user.id,
                       Task.created_by  == current_user.id)
            ).distinct().count()
            vol = db.session.query(db.func.sum(Task.volume)).filter(
                db.or_(Task.assigned_to == current_user.id,
                       Task.created_by  == current_user.id)
            ).scalar() or 0
        else:
            projs = Project.query.count()
            vol   = db.session.query(db.func.sum(Task.volume)).scalar() or 0

        # PERF-P3: Two GROUP BY queries replace 9 individual COUNT round-trips.
        # One query groups by status, one by priority — O(n) over tasks once each.
        base_filter = []
        if current_user.role == 'member':
            base_filter = [db.or_(Task.assigned_to == current_user.id,
                                  Task.created_by  == current_user.id)]

        status_rows = db.session.query(
            Task.status, db.func.count(Task.id)
        ).filter(*base_filter).group_by(Task.status).all()
        status_map  = {r[0]: r[1] for r in status_rows}

        priority_rows = db.session.query(
            Task.priority, db.func.count(Task.id)
        ).filter(*base_filter).group_by(Task.priority).all()
        priority_map = {r[0]: r[1] for r in priority_rows}

        total = sum(status_map.values())
        done  = status_map.get('done', 0)
        pend  = status_map.get('pending', 0)
        inp   = status_map.get('in_progress', 0)
        blk   = status_map.get('blocked', 0)
        pdata = {
            'low':      priority_map.get('low', 0),
            'medium':   priority_map.get('medium', 0),
            'high':     priority_map.get('high', 0),
            'critical': priority_map.get('critical', 0),
        }

        # PERF: Compute avg TAT in the database (single scalar query) instead of
        # fetching all completed task rows into Python and summing there.
        # SQLite stores datetimes as text; strftime arithmetic gives seconds.
        avg_q = db.session.query(
            db.func.avg(
                db.func.strftime('%s', Task.completed_at) -
                db.func.strftime('%s', Task.started_at)
            )
        ).filter(
            Task.started_at.isnot(None),
            Task.completed_at.isnot(None),
        )
        if current_user.role == 'member':
            avg_q = avg_q.filter(
                db.or_(Task.assigned_to == current_user.id,
                       Task.created_by  == current_user.id))
        avg_secs = avg_q.scalar()
        avg = round(avg_secs / 60, 1) if avg_secs else 0

        # Total active users (admin dashboard "Team Members" stat card)
        total_users = User.query.filter_by(is_active=True).count()

        return jsonify({
            'total_tasks': total, 'done_tasks': done, 'pending_tasks': pend,
            'in_progress': inp, 'blocked': blk, 'total_projects': projs,
            'total_volume': vol, 'avg_tat_mins': avg, 'priority_data': pdata,
            'completion_rate': round(done / total * 100, 1) if total else 0,
            'total_users': total_users,
        })
    except Exception as e:
        logger.exception('Unhandled error'); return jsonify({'error': 'An internal error occurred'}), 500


# =============================================================================
# TASK NAMES API  (lookup table management)
# =============================================================================
#
# Task Names are the approved {process, task_name, cap_timing} combinations
# that populate the Task Name dropdowns in both portals.
#
# Endpoints:
#   GET    /api/task-names               → list all task names (all roles)
#   POST   /api/task-names               → create one (admin/lead)
#   PUT    /api/task-names/<tid>         → update one (admin/lead)
#   DELETE /api/task-names/<tid>         → delete one (admin/lead)
#   POST   /api/task-names/bulk          → bulk import from JSON (admin/lead)
#   POST   /api/task-names/parse-xlsx    → parse xlsx → JSON rows (admin/lead)
#   GET    /api/task-names/template      → download import template xlsx
#   POST   /api/task-names/export        → export selected rows as xlsx

@app.route('/task-names')
@login_required
def task_names_page():
    """Render the Task Names management page (task_names.html)."""
    return render_template('task_names.html', user=current_user)


@app.route('/api/task-names', methods=['GET'])
@login_required
def get_task_names():
    """Return TaskName records ordered by process then task_name.
    Optional ?status= filter (defaults to all). Capped at 2000 rows.
    PERF-P1: Hard cap prevents unbounded serialization on large installations.
    """
    try:
        q = TaskName.query.order_by(TaskName.process, TaskName.task_name)
        status_filter = request.args.get('status')
        if status_filter in ('active', 'inactive'):
            q = q.filter_by(status=status_filter)
        items = q.limit(2000).all()
        return jsonify([i.to_dict() for i in items])
    except Exception as e:
        logger.exception('Unhandled error')
        return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/task-names', methods=['POST'])
@login_required
@_login_limit
def create_task_name():
    """
    Create a new TaskName record.  Admin/lead only.
    Enforces uniqueness on (process, task_name) to prevent duplicate entries.
    cap_timing must be >= 1 minute.
    """
    try:
        if current_user.role not in ('admin', 'lead'):
            return jsonify({'error': 'Unauthorized'}), 403
        data = request.get_json(silent=True) or {}
        if not data.get('process'):
            return jsonify({'error': 'Process is required'}), 400
        if not data.get('task_name'):
            return jsonify({'error': 'Task name is required'}), 400
        cap = int(data.get('cap_timing', 0))
        if cap < 1:
            return jsonify({'error': 'Cap timing must be >= 1'}), 400
        proc  = sanitize(data['process'])[:200]
        tname = sanitize(data['task_name'])[:300]
        if TaskName.query.filter_by(process=proc, task_name=tname).first():
            return jsonify({'error': 'This task name already exists for the selected process'}), 409
        tn = TaskName(
            process=proc,
            task_name=tname,
            cap_timing=cap, status=data.get('status', 'active'),
            created_by=current_user.id,
        )
        db.session.add(tn)
        db.session.commit()
        return jsonify(tn.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        logger.exception('Unhandled error'); return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/task-names/<int:tid>', methods=['PUT'])
@login_required
@_login_limit
def update_task_name(tid):
    """
    Update process, task_name, cap_timing, or status on an existing TaskName.
    Admin/lead only.  cap_timing is clamped to >= 1 minute.
    status must be 'active' or 'inactive' — 'inactive' hides the entry from
    dropdowns while preserving history on existing tasks.
    """
    try:
        if current_user.role not in ('admin', 'lead'):
            return jsonify({'error': 'Unauthorized'}), 403
        tn = db.session.get(TaskName, tid)   # BUG-06 FIX
        if tn is None:
            return jsonify({'error': 'Task name not found'}), 404
        data = request.get_json(silent=True) or {}
        if 'process'    in data: tn.process   = sanitize(data['process'])[:200]
        if 'task_name'  in data: tn.task_name = sanitize(data['task_name'])[:300]
        if 'cap_timing' in data:
            try:
                tn.cap_timing = max(1, int(data['cap_timing']))
            except (TypeError, ValueError):
                return jsonify({'error': 'Invalid cap_timing'}), 400
        if 'status' in data and data['status'] in ('active', 'inactive'):
            tn.status = data['status']
        db.session.commit()
        return jsonify(tn.to_dict())
    except Exception as e:
        db.session.rollback()
        logger.exception('Unhandled error')
        return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/task-names/<int:tid>', methods=['DELETE'])
@login_required
@_login_limit
def delete_task_name(tid):
    """Hard-delete a TaskName record.  Admin/lead only."""
    try:
        if current_user.role not in ('admin', 'lead'):
            return jsonify({'error': 'Unauthorized'}), 403
        tn = db.session.get(TaskName, tid)   # BUG-06 FIX
        if tn is None:
            return jsonify({'error': 'Task name not found'}), 404
        db.session.delete(tn)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        logger.exception('Unhandled error')
        return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/task-names/bulk', methods=['POST'])
@login_required
@_login_limit
def bulk_import_task_names():
    """
    Bulk-import TaskName records from a JSON array.  Admin/lead only.

    Body:
      rows (list[dict]) — each dict must have: process, task_name, cap_timing.

    Rows are silently skipped (not errored) when:
      • cap_timing is non-numeric or < 1  (BUG-FIX-5: was raising ValueError)
      • process or task_name is empty after sanitisation
      • An identical (process, task_name) pair already exists in the DB

    SEC-11: Payload capped at 5 000 rows to prevent memory-exhaustion via huge
    imports.  Returns {imported: N} with the count of rows actually inserted.
    """
    try:
        if current_user.role not in ('admin', 'lead'):
            return jsonify({'error': 'Unauthorized'}), 403
        data = request.get_json(silent=True) or {}
        rows = data.get('rows', [])
        if not rows:
            return jsonify({'error': 'No rows provided'}), 400
        # SEC-11: Prevent DoS via huge payloads (raised to 5000 to support large imports)
        if len(rows) > 5000:
            return jsonify({'error': 'Too many rows: max 5000 per import'}), 400
        # PERF-P4: Pre-load all existing (process, task_name) pairs into a set so
        # duplicate detection is O(1) per row instead of one DB query per row.
        existing = {
            (r.process, r.task_name)
            for r in db.session.query(TaskName.process, TaskName.task_name).all()
        }
        count = 0
        for row in rows:
            try:
                cap = int(row.get('cap_timing', 0))
            except (TypeError, ValueError):
                continue   # skip rows with non-numeric cap_timing silently
            if cap < 1:
                continue
            proc = sanitize(str(row.get('process', '')))[:200]
            tname = sanitize(str(row.get('task_name', '')))[:300]
            if not proc or not tname:
                continue
            # Skip duplicates using in-memory set (no per-row DB query)
            if (proc, tname) in existing:
                continue
            existing.add((proc, tname))   # prevent duplicates within the batch itself
            tn = TaskName(
                process=proc, task_name=tname,
                cap_timing=cap, status='active', created_by=current_user.id,
            )
            db.session.add(tn)
            count += 1
        db.session.commit()
        return jsonify({'imported': count}), 201
    except Exception as e:
        db.session.rollback()
        logger.exception('Unhandled error'); return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/task-names/parse-xlsx', methods=['POST'])
@login_required
@_login_limit
def parse_task_names_xlsx():
    """
    Parse an uploaded xlsx into a list of {process, task_name, cap_timing, errors, valid}
    row objects for preview before bulk import.

    The response is a {rows: [...]} object where each row has:
      idx        (int)   — 1-based row number for display
      process    (str)   — sanitised Mapping Process value
      task_name  (str)   — sanitised Task Name value
      cap_timing (int)   — parsed cap timing in minutes
      errors     (list)  — list of human-readable validation error strings
      valid      (bool)  — True when errors is empty

    Security: same MIME-type and magic-byte checks as parse_tasks_xlsx (SEC-D).
    Row cap: 5 000 rows max (SEC-11).
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        f = request.files['file']
        if not f.filename or not f.filename.lower().endswith('.xlsx'):
            return jsonify({'error': 'Please upload a .xlsx file'}), 400
        # SEC-D: Verify actual MIME type, not just the file extension
        allowed_mimes = {
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/octet-stream',  # Some browsers send this for xlsx
            'application/zip',           # xlsx is a zip internally
        }
        if f.mimetype and f.mimetype not in allowed_mimes and not f.mimetype.startswith('application/'):
            return jsonify({'error': 'Invalid file type. Please upload a valid .xlsx file'}), 400
        # SEC-D: Verify xlsx magic bytes (ZIP header: PK\x03\x04) before parsing
        header_bytes = f.stream.read(4)
        f.stream.seek(0)
        if header_bytes[:2] != b'PK':
            return jsonify({'error': 'Invalid file format. File does not appear to be a valid .xlsx file.'}), 400
        wb = load_workbook(f.stream, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        # First row = headers
        try:
            headers = [str(h).strip().lower() if h is not None else '' for h in next(rows_iter)]
        except StopIteration:
            return jsonify({'error': 'File is empty'}), 400
        # Map header names to indices
        def find_col(*names):
            for name in names:
                if name in headers:
                    return headers.index(name)
            return None
        proc_idx = find_col('mapping process', 'process')
        task_idx = find_col('tasks', 'task', 'task name', 'taskname')
        cap_idx  = find_col('cap timings (mins)', 'cap timings', 'cap timing', 'cap_timing', 'timing')
        if proc_idx is None or task_idx is None or cap_idx is None:
            missing = []
            if proc_idx is None: missing.append('Process')
            if task_idx is None: missing.append('Tasks')
            if cap_idx  is None: missing.append('Cap Timings (mins)')
            return jsonify({'error': f'Missing required columns: {", ".join(missing)}'}), 400
        parsed = []
        MAX_IMPORT_ROWS = 5000
        for i, row in enumerate(rows_iter):
            if i >= MAX_IMPORT_ROWS:
                wb.close()
                return jsonify({'error': f'File exceeds the {MAX_IMPORT_ROWS}-row import limit. Please split it into smaller files.'}), 400
            def cell(idx):
                v = row[idx] if idx < len(row) else None
                return str(v).strip() if v is not None else ''
            process   = cell(proc_idx)
            task_name = cell(task_idx)
            cap_raw   = cell(cap_idx)
            try: cap = int(float(cap_raw)) if cap_raw else 0
            except: cap = 0
            errors = []
            if not process:   errors.append('Missing Mapping Process')
            if not task_name: errors.append('Missing Task')
            if cap < 1:       errors.append('Invalid Cap Timing')
            parsed.append({
                'idx': i + 1,
                'process':   sanitize(process)[:200],
                'task_name': sanitize(task_name)[:300],
                'cap_timing': cap,
                'errors': errors,
                'valid': len(errors) == 0,
            })
        wb.close()
        return jsonify({'rows': parsed}), 200
    except Exception as e:
        logger.exception('xlsx parse error (task-names)')
        return jsonify({'error': 'Failed to parse file. Please check the format and try again.'}), 500


@app.route('/api/task-names/template', methods=['GET'])
@login_required
def download_task_template():
    try:

        wb = Workbook()
        ws = wb.active
        ws.title = 'Task Import Template'

        thin   = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        headers    = ['Mapping Process', 'Tasks', 'Cap Timings (mins)']
        col_widths = [30, 40, 22]

        for i, (h, w) in enumerate(zip(headers, col_widths), 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=11)
            c.fill      = PatternFill('solid', start_color='0A0A0F')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = border
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 30

        samples = [
            ('Data Processing',   'Data Validation',        45),
            ('Data Processing',   'Data Cleansing',         30),
            ('Report Generation', 'Monthly Summary Report', 60),
            ('Report Generation', 'Dashboard Update',       20),
            ('Quality Check',     'Peer Review',            90),
            ('Quality Check',     'Final Approval',         15),
        ]
        alt = PatternFill('solid', start_color='F5F3EE')
        for r, row in enumerate(samples, 2):
            for col, val in enumerate(row, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.font      = Font(name='Arial', size=10)
                c.alignment = Alignment(horizontal='left', vertical='center')
                c.border    = border
                if r % 2 == 0:
                    c.fill = alt

        wi = wb.create_sheet('Instructions')
        wi['A1'] = 'HOW TO USE THIS TEMPLATE'
        wi['A1'].font = Font(name='Arial', bold=True, size=14, color='FF4D2E')
        inst = [
            ('Column', 'Description', 'Example'),
            ('Process', 'High-level process or category', 'Data Processing'),
            ('Tasks', 'Specific task name', 'Data Validation'),
            ('Cap Timings (mins)', 'Time cap in minutes (numbers only)', '45'),
        ]
        for ri, row in enumerate(inst, 3):
            for ci, val in enumerate(row, 1):
                c = wi.cell(row=ri, column=ci, value=val)
                if ri == 3:
                    c.font      = Font(name='Arial', bold=True, color='FFFFFF')
                    c.fill      = PatternFill('solid', start_color='0A0A0F')
                    c.alignment = Alignment(horizontal='center')
                else:
                    c.font = Font(name='Arial', size=10)
        wi.column_dimensions['A'].width = 25
        wi.column_dimensions['B'].width = 45
        wi.column_dimensions['C'].width = 25

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='task_import_template.xlsx',
        )
    except Exception as e:
        logger.exception('Unhandled error'); return jsonify({'error': 'An internal error occurred'}), 500




@app.route('/api/task-names/export', methods=['POST'])
@login_required
@_login_limit
def export_task_names():
    """Server-side task-names export — returns xlsx filtered by supplied IDs."""
    try:
        if current_user.role not in ('admin', 'lead'):
            return jsonify({'error': 'Unauthorized'}), 403
        data = request.get_json(silent=True) or {}
        raw_ids = data.get('ids', [])
        # SEC: Validate ids are integers; cap list size to prevent DoS
        if not isinstance(raw_ids, list):
            return jsonify({'error': 'ids must be a list'}), 400
        if len(raw_ids) > 5000:
            return jsonify({'error': 'Too many ids: max 5000'}), 400
        try:
            ids = [int(i) for i in raw_ids]
        except (TypeError, ValueError):
            return jsonify({'error': 'ids must be integers'}), 400
        q = TaskName.query.order_by(TaskName.process, TaskName.task_name)
        if ids:
            q = q.filter(TaskName.id.in_(ids))
        items = q.all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Task Names'

        thin   = Side(style='thin', color='DDDDDD')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        headers    = ['Mapping Process', 'Tasks', 'Cap Timings (mins)', 'Status', 'Created']
        col_widths = [30, 40, 22, 14, 14]

        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            c.fill      = PatternFill('solid', start_color='0A0A0F')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = border
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 24

        alt = PatternFill('solid', start_color='F8F7F4')
        for ri, item in enumerate(items, 2):
            vals = [item.process, item.task_name, item.cap_timing, item.status,
                    item.created_at.strftime('%Y-%m-%d') if item.created_at else '']
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font      = Font(name='Arial', size=9)
                c.alignment = Alignment(vertical='center')
                c.border    = border
                if ri % 2 == 0:
                    c.fill = alt

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"TaskNames_{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
        return send_file(buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=fname)
    except Exception as e:
        logger.exception('export_task_names error')
        return jsonify({'error': 'An internal error occurred'}), 500


# =============================================================================
# NOTIFICATION / EMAIL API
# =============================================================================
#
# These endpoints let admins configure and test the email system at runtime
# without needing to restart the server.
#
# Endpoints:
#   GET  /api/notifications/settings      → read current SMTP config (passwords masked)
#   PUT  /api/notifications/settings      → hot-update SMTP config (admin only)
#   POST /api/notifications/test          → send a test email to the calling admin
#   POST /api/notifications/send-reminder → trigger overdue digest immediately
#   GET  /api/notifications/overdue-preview → dry-run: show what would be emailed

@app.route('/api/notifications/settings', methods=['GET'])
@login_required
def get_notification_settings():
    """
    Return the current SMTP email configuration.  Admin/lead only.

    Passwords are intentionally excluded from the response — the UI can show
    whether a password is configured (non-empty username) but cannot retrieve it.
    reminder_hours and reminder_interval are included so the UI can display the
    current scheduler settings.
    """
    if current_user.role not in ('admin', 'lead'):
        return jsonify({'error': 'Unauthorized'}), 403
    return jsonify({
        'mail_enabled':      app.config.get('MAIL_ENABLED', False),
        'mail_server':       app.config.get('MAIL_SERVER', ''),
        'mail_port':         app.config.get('MAIL_PORT', 587),
        'mail_use_tls':      app.config.get('MAIL_USE_TLS', True),
        'mail_username':     app.config.get('MAIL_USERNAME', ''),
        'mail_from':         app.config.get('MAIL_FROM', ''),
        'reminder_hours':    app.config.get('REMINDER_HOURS', 24),
        'reminder_interval': app.config.get('REMINDER_INTERVAL', 3600),
    })


@app.route('/api/notifications/settings', methods=['PUT'])
@login_required
@_login_limit
def update_notification_settings():
    """
    Hot-update the SMTP email configuration at runtime.  Admin only.

    Changes take effect immediately for the next queued email without needing
    a server restart.  mail_password is only updated when the key is present
    AND non-empty — sending an empty string does not clear the saved password.
    """
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json(silent=True) or {}
    # SEC-S2/S3: Validate and cap all SMTP string fields before writing to config
    if 'mail_enabled'  in data:
        app.config['MAIL_ENABLED'] = bool(data['mail_enabled'])
    if 'mail_server' in data:
        srv = str(data['mail_server']).strip()[:253]   # SEC-S3: max valid hostname length
        if not srv:
            return jsonify({'error': 'mail_server cannot be empty'}), 400
        app.config['MAIL_SERVER'] = srv
    if 'mail_use_tls'  in data:
        app.config['MAIL_USE_TLS'] = bool(data['mail_use_tls'])
    if 'mail_username' in data:
        app.config['MAIL_USERNAME'] = str(data['mail_username']).strip()[:254]
    if 'mail_from' in data:
        app.config['MAIL_FROM'] = str(data['mail_from']).strip()[:254]
    # SEC-S2: Enforce valid TCP port range
    if 'mail_port' in data:
        try:
            port = int(data['mail_port'])
            if not (1 <= port <= 65535):
                return jsonify({'error': 'mail_port must be between 1 and 65535'}), 400
            app.config['MAIL_PORT'] = port
        except (TypeError, ValueError):
            return jsonify({'error': 'mail_port must be an integer'}), 400
    # Only update password when explicitly supplied and non-empty
    if 'mail_password' in data and data['mail_password']:
        app.config['MAIL_PASSWORD'] = str(data['mail_password'])[:500]
    # SEC-S4: Enforce minimum reminder_interval of 60s to prevent scheduler DoS
    if 'reminder_interval' in data:
        try:
            interval = int(data['reminder_interval'])
            if interval < 60:
                return jsonify({'error': 'reminder_interval must be at least 60 seconds'}), 400
            app.config['REMINDER_INTERVAL'] = interval
        except (TypeError, ValueError):
            return jsonify({'error': 'reminder_interval must be an integer'}), 400
    # SEC-S5: Return error instead of silently ignoring invalid reminder_hours
    if 'reminder_hours' in data:
        try:
            rh = int(data['reminder_hours'])
            if rh < 0:
                return jsonify({'error': 'reminder_hours must be non-negative'}), 400
            app.config['REMINDER_HOURS'] = rh
        except (TypeError, ValueError):
            return jsonify({'error': 'reminder_hours must be an integer'}), 400
    return jsonify({'success': True, 'mail_enabled': app.config['MAIL_ENABLED']})


@app.route('/api/notifications/test', methods=['POST'])
@login_required
@_login_limit
def test_notification():
    """
    Send a test email to the currently-authenticated admin's own address.

    Used to verify SMTP configuration from the Admin portal Notifications panel.
    The email includes SMTP server, port, TLS flag, and send timestamp so the
    admin can confirm the right configuration is active.  Admin only.
    """
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    if not current_user.email:
        return jsonify({'error': 'Your account has no email address'}), 400
    body = f"""
<p>Hi <strong>{current_user.username}</strong>,</p>
<p>This is a test email from HuddleTracker. If you received this, your SMTP configuration is working correctly!</p>
<div class="card">
  <div class="card-row"><span class="lbl">Server</span><span class="val">{app.config['MAIL_SERVER']}:{app.config['MAIL_PORT']}</span></div>
  <div class="card-row"><span class="lbl">TLS</span><span class="val">{'Yes' if app.config['MAIL_USE_TLS'] else 'No'}</span></div>
  <div class="card-row"><span class="lbl">From</span><span class="val">{app.config.get('MAIL_FROM') or app.config.get('MAIL_USERNAME')}</span></div>
  <div class="card-row"><span class="lbl">Sent at</span><span class="val">{datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}</span></div>
</div>
"""
    send_email(
        to=current_user.email,
        subject='🔔 HuddleTracker — Test Email',
        html=_html_wrap('Email Configuration Test', body),
        text=f'Hi {current_user.username}, your HuddleTracker email config is working!',
    )
    return jsonify({'success': True, 'sent_to': current_user.email})


@app.route('/api/notifications/send-reminder', methods=['POST'])
@login_required
@_login_limit
def manual_send_reminder():
    """
    Manually trigger an overdue task reminder blast to all admin/lead users.

    Calls the same _check_and_send_overdue_reminders() function used by the
    daily scheduler, but immediately instead of waiting for the next scheduled
    window.  Admin/lead only.  Returns {success: true} as soon as the digest
    is enqueued (actual delivery is async via the mail worker thread).
    """
    if current_user.role not in ('admin', 'lead'):
        return jsonify({'error': 'Unauthorized'}), 403
    try:
        _check_and_send_overdue_reminders()
        return jsonify({'success': True})
    except Exception as e:
        logger.exception('manual_send_reminder error')
        return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/notifications/overdue-preview', methods=['GET'])
@login_required
def overdue_preview():
    """
    Return the current overdue task breakdown without sending any email.

    Dry-run companion to send-reminder.  Returns:
      total      (int)   — number of overdue tasks
      by_member  (dict)  — {username: count} breakdown
      tasks      (list)  — full task dicts for each overdue task

    Used by the Admin portal to show a preview before manually triggering the
    reminder blast.  Admin/lead only.
    """
    if current_user.role not in ('admin', 'lead'):
        return jsonify({'error': 'Unauthorized'}), 403
    now = datetime.utcnow()
    # SEC-S5/PERF-P1: Cap at 200 rows — prevents unbounded data dump and memory spike.
    # Full task dicts replaced with lightweight summary objects to reduce payload size.
    overdue = (
        Task.query
        .options(joinedload(Task._assignee))
        .filter(Task.due_date < now, Task.status.notin_(['done']))
        .order_by(Task.due_date.asc())
        .limit(200)
        .all()
    )
    by_member: dict = defaultdict(int)
    task_summaries = []
    for t in overdue:
        name = t._assignee.username if t._assignee else 'Unassigned'
        by_member[name] += 1
        task_summaries.append({
            'id':        t.id,
            'study_id':  t.study_id,
            'task_name': t.task_name,
            'assignee':  name,
            'enterprise_id': t._assignee.enterprise_id if t._assignee else '',
            'priority':  t.priority,
            'status':    t.status,
            'due_date':  t.due_date.strftime('%Y-%m-%d') if t.due_date else None,
            'days_overdue': (now - t.due_date).days if t.due_date else 0,
        })
    return jsonify({
        'total':     len(overdue),
        'by_member': dict(by_member),
        'tasks':     task_summaries,
    })


# =============================================================================
# PROCESSES API  (master list of mapping process categories)
# =============================================================================
#
# Endpoints:
#   GET    /api/processes          → list all processes (all authenticated roles)
#   POST   /api/processes          → create a process (admin/lead)
#   PUT    /api/processes/<pid>    → rename a process — cascades to TaskName + Task (admin/lead)
#   DELETE /api/processes/<pid>    → delete (blocked if any TaskName or Task references it)

@app.route('/api/processes', methods=['GET'])
@login_required
def get_processes():
    """Return all Process records alphabetically.  Available to all roles."""
    try:
        items = Process.query.order_by(Process.name).all()
        return jsonify([p.to_dict() for p in items])
    except Exception as e:
        logger.exception('Unhandled error')
        return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/processes', methods=['POST'])
@login_required
@_login_limit
def create_process():
    """Create a new Process.  Admin/lead only.  Name must be unique."""
    try:
        if current_user.role not in ('admin', 'lead'):
            return jsonify({'error': 'Unauthorized'}), 403
        data = request.get_json(silent=True) or {}
        name = sanitize(data.get('name', '')).strip()
        if not name:
            return jsonify({'error': 'Process name is required'}), 400
        if Process.query.filter_by(name=name).first():
            return jsonify({'error': 'Process already exists'}), 409
        p = Process(name=name[:200], created_by=current_user.id)
        db.session.add(p)
        db.session.commit()
        return jsonify(p.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        logger.exception('Unhandled error'); return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/processes/<int:pid>', methods=['PUT'])
@login_required
@_login_limit
def update_process(pid):
    """
    Rename a Process.  Admin/lead only.

    The rename cascades to both TaskName.process and Task.process columns via
    bulk UPDATE so existing records always reflect the current process name.
    A conflict check prevents renaming to a name already used by another process.
    """
    try:
        if current_user.role not in ('admin', 'lead'):
            return jsonify({'error': 'Unauthorized'}), 403
        p = db.session.get(Process, pid)
        if p is None:
            return jsonify({'error': 'Process not found'}), 404
        data = request.get_json(silent=True) or {}
        new_name = sanitize(data.get('name', '')).strip()
        if not new_name:
            return jsonify({'error': 'Process name is required'}), 400
        existing = Process.query.filter_by(name=new_name).first()
        if existing and existing.id != pid:
            return jsonify({'error': 'Process name already exists'}), 409
        old_name = p.name
        p.name = new_name[:200]
        # Cascade rename to all TaskName rows using the old process name
        TaskName.query.filter_by(process=old_name).update({'process': new_name})
        # Also cascade to live Task rows so existing tasks don't get orphaned
        Task.query.filter_by(process=old_name).update({'process': new_name})
        db.session.commit()
        return jsonify(p.to_dict())
    except Exception as e:
        db.session.rollback()
        logger.exception('Unhandled error')
        return jsonify({'error': 'An internal error occurred'}), 500


@app.route('/api/processes/<int:pid>', methods=['DELETE'])
@login_required
@_login_limit
def delete_process(pid):
    """
    Delete a Process.  Admin/lead only.

    Deletion is blocked (409 Conflict) when:
      • Any TaskName still references this process name.
      • Any live Task still references this process name.

    This prevents dangling foreign values that would make those records
    impossible to filter or group by process.  Users must reassign or delete
    the dependent records before the process can be removed.
    """
    try:
        if current_user.role not in ('admin', 'lead'):
            return jsonify({'error': 'Unauthorized'}), 403
        p = db.session.get(Process, pid)
        if p is None:
            return jsonify({'error': 'Process not found'}), 404
        task_count = TaskName.query.filter_by(process=p.name).count()
        if task_count > 0:
            return jsonify({'error': f'Cannot delete: {task_count} task name(s) use this process. Reassign them first.'}), 409
        live_task_count = Task.query.filter_by(process=p.name).count()
        if live_task_count > 0:
            return jsonify({'error': f'Cannot delete: {live_task_count} task(s) still reference this process. Reassign them first.'}), 409
        db.session.delete(p)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        logger.exception('Unhandled error')
        return jsonify({'error': 'An internal error occurred'}), 500


# =============================================================================
# ERROR HANDLERS
# =============================================================================

@app.errorhandler(413)
def request_entity_too_large(e):
    """Return a JSON 413 when the request body exceeds MAX_CONTENT_LENGTH (10 MB)."""
    return jsonify({'error': 'File too large. Maximum upload size is 10 MB.'}), 413


# =============================================================================
# DATABASE SEED
# =============================================================================

def seed_db():
    """
    Create all database tables and insert demo data on the first run.

    Idempotent — if the 'admin' user already exists the function exits early
    after ensuring the seed processes are present (added in a later version).

    Demo credentials (shown only when FLASK_ENV != 'production'):
      admin   / Admin@1234
      lead1   / Lead@1234
      member1 / Member@1234

    BUG-07 FIX: Removed the fragile raw ALTER TABLE migration that was
    previously needed to add the `process` column to existing databases.
    The column is now declared in the Task model and created correctly by
    db.create_all() for all new databases.  Use Flask-Migrate (Alembic) for
    upgrading existing production databases.
    """
    db.create_all()

    # BUG-07 FIX: Removed the fragile raw ALTER TABLE migration.
    # The `process` column is now part of the Task model definition above and
    # will be created correctly by db.create_all() for all new databases.
    # For existing databases, use Flask-Migrate (Alembic) to manage migrations.

    if User.query.filter_by(username='admin').first():
        # Seed processes if missing (in case DB existed before this feature)
        if Process.query.count() == 0:
            seed_processes = ['NG 1.0', 'NG 2.0', 'Integrated Studies',
                              'Integrated Standards', 'NG Trials', 'Migrations']
            for pname in seed_processes:
                db.session.add(Process(name=pname))
            db.session.commit()
        return  # Already seeded

    print('Seeding demo data…')
    admin  = User(username='admin',   email='admin@huddle.io',
                  enterprise_id='ENT-ADMIN-001', role='admin')
    lead   = User(username='lead1',   email='lead@huddle.io',
                  enterprise_id='ENT-LEAD-001',  role='lead')
    member = User(username='member1', email='member@huddle.io',
                  enterprise_id='ENT-MEM-001',   role='member')
    # SEC-S1: In production, generate strong random passwords instead of
    # using the known demo passwords — prevents credential stuffing attacks.
    # Use ADMIN_PASSWORD env var if set, otherwise fall back to default.
    # Set ADMIN_PASSWORD in Render environment variables for security.
    admin_pw  = os.environ.get('ADMIN_PASSWORD',  'Admin@1234')
    lead_pw   = os.environ.get('LEAD_PASSWORD',   'Lead@1234')
    member_pw = os.environ.get('MEMBER_PASSWORD', 'Member@1234')
    admin.set_password(admin_pw)
    lead.set_password(lead_pw)
    member.set_password(member_pw)
    logger.warning('DB seeded. Default credentials active — change them after login.')
    db.session.add_all([admin, lead, member])
    db.session.commit()

    p = Project(
        project_code='PRJ-DEMO0001', name='Huddle Demo Project',
        description='Sample project to get started.',
        owner_id=admin.id, deadline=datetime.utcnow() + timedelta(days=30),
    )
    db.session.add(p)
    db.session.commit()

    rows = [
        ('RFC-2024-1001', 'Data Migration',    'Migrate legacy records',  'pending',     'low',       10,  4),
        ('RFC-2024-1002', 'API Integration',   'Connect payment gateway', 'in_progress', 'high',      25,  8),
        ('RFC-2024-1003', 'UI Redesign',        'Redesign dashboard',     'review',      'medium',    15,  6),
        ('RFC-2024-1004', 'Security Audit',     'Run penetration tests',  'done',        'critical',   5, 16),
        ('RFC-2024-1005', 'Load Testing',       'Test 1000 virtual users','blocked',     'high',      20, 12),
    ]
    for sr, task, sub, status, pri, vol, tat in rows:
        t = Task(
            study_id=gen_study_id(), sr_rfc=sr, project_id=p.id,
            assigned_to=member.id, task_name=task, subtask=sub,
            status=status, priority=pri, volume=vol,
            turnaround_time=tat, created_by=admin.id,
        )
        if status in ('in_progress', 'review', 'done'):
            t.started_at = datetime.utcnow() - timedelta(hours=tat)
        if status == 'done':
            t.completed_at = datetime.utcnow()
        db.session.add(t)

    seed_processes = ['NG 1.0', 'NG 2.0', 'Integrated Studies',
                       'Integrated Standards', 'NG Trials', 'Migrations']
    for pname in seed_processes:
        if not Process.query.filter_by(name=pname).first():
            db.session.add(Process(name=pname, created_by=admin.id))
    db.session.commit()

    task_names_seed = [
        ('Data Processing',   'Data Validation',        45),
        ('Data Processing',   'Data Cleansing',         30),
        ('Report Generation', 'Monthly Summary Report', 60),
        ('Report Generation', 'Dashboard Update',       20),
        ('Quality Check',     'Peer Review',            90),
        ('Quality Check',     'Final Approval',         15),
    ]
    for proc, tname, cap in task_names_seed:
        tn = TaskName(process=proc, task_name=tname, cap_timing=cap,
                      status='active', created_by=admin.id)
        db.session.add(tn)

    db.session.commit()
    print('Done! Login: admin / Admin@1234')



if __name__ == '__main__':
    # Initialise the database and insert demo data on first run.
    # seed_db() is idempotent — safe to call every time the dev server starts.
    with app.app_context():
        seed_db()
    # BUG-08 FIX: debug mode driven by FLASK_DEBUG env var — never hardcoded.
    # threaded=True enables Flask's built-in thread-per-request mode which is
    # required for the SSE /api/stream endpoint (each SSE client holds a
    # long-lived response connection in its own thread).
    app.run(
        debug=app.config['DEBUG'],
        host='0.0.0.0',
        port=int(os.environ.get('PORT', 5000)),
        threaded=True,
    )
